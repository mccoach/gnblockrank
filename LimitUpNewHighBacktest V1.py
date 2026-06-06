import os
import re
import sys
import glob
import json
import struct
import queue
import ntpath
import threading
import calendar
from copy import deepcopy
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

# ============================================================
# 常量
# ============================================================

CONFIG_FILE_NAME = "LimitUpNewHighBacktest_config.json"
ICON_FILE_NAME = "icon.ico"

DEFAULT_TIMEZONE_NAME = "Asia/Shanghai"

TDX_DAY_RECORD_SIZE = 32
TDX_TNF_HEADER_SIZE = 50
TDX_TNF_RECORD_SIZE = 360
PRICE_DIVISOR = 100.0

MARKET_DAY_PREFIX = {
    "SH": "sh",
    "SZ": "sz",
}

TNF_FILES = {
    "SH": "shs.tnf",
    "SZ": "szs.tnf",
}

DEFAULT_LIMIT_UP_THRESHOLD = 19.8
DEFAULT_EXCLUDE_LOOKBACK_DAYS = 10
DEFAULT_EXCLUDE_PCT_THRESHOLD = 18.0
DEFAULT_AUTO_OPEN_FILE = True

UI_BUTTON_WIDTH = 10
UI_MAIN_BUTTON_WIDTH = 14

# ============================================================
# 路径工具
# ============================================================


def normalize_path_slashes(path):
    if path is None:
        return ""
    return str(path).replace("/", "\\")


def strip_outer_path_quotes(path):
    text = str(path or "").strip()
    pairs = [
        ('"', '"'),
        ("'", "'"),
        ("“", "”"),
        ("‘", "’"),
    ]

    changed = True
    while changed and len(text) >= 2:
        changed = False
        for left, right in pairs:
            if text.startswith(left) and text.endswith(right):
                text = text[1:-1].strip()
                changed = True
                break

    return text


def is_windows_root_path(path):
    path = normalize_path_slashes(path).strip()
    if not path:
        return False

    drive, tail = ntpath.splitdrive(path)
    if drive and tail == "\\":
        return True

    if path.startswith("\\\\"):
        parts = [p for p in path.split("\\") if p]
        return len(parts) == 2

    return False


def strip_redundant_trailing_slashes(path):
    text = normalize_path_slashes(path).strip()

    while len(text) > 1 and text.endswith(
            "\\") and not is_windows_root_path(text):
        text = text[:-1]

    return text


def normalize_user_path(path, trim_trailing_slash=True):
    text = strip_outer_path_quotes(path)
    text = normalize_path_slashes(text).strip()

    if trim_trailing_slash:
        text = strip_redundant_trailing_slashes(text)

    return text


def win_join(*parts):
    return normalize_path_slashes(ntpath.join(*parts))


def win_dirname(path):
    return normalize_path_slashes(ntpath.dirname(path))


def get_program_dir():
    if getattr(sys, "frozen", False):
        return normalize_path_slashes(os.path.dirname(sys.executable))
    return normalize_path_slashes(os.path.dirname(os.path.abspath(__file__)))


def get_resource_path(filename):
    if getattr(sys, "frozen", False):
        base_dir = getattr(sys, "_MEIPASS", get_program_dir())
    else:
        base_dir = get_program_dir()

    return normalize_path_slashes(os.path.join(base_dir, filename))


def get_config_path():
    return win_join(get_program_dir(), CONFIG_FILE_NAME)


def existing_path_for_root_search(path):
    path = normalize_user_path(path, True)
    if not path:
        return ""

    if os.path.isfile(path):
        return win_dirname(path)

    return path


def iter_parent_dirs(path):
    current = existing_path_for_root_search(path)

    while current:
        yield current

        parent = win_dirname(current)
        if not parent or parent == current:
            break

        current = parent


# ============================================================
# 日期工具
# ============================================================


def one_month_before(d):
    year = d.year
    month = d.month - 1

    if month == 0:
        year -= 1
        month = 12

    day = min(d.day, calendar.monthrange(year, month)[1])
    return d.replace(year=year, month=month, day=day)


def default_start_end_dates():
    today = datetime.now(ZoneInfo(DEFAULT_TIMEZONE_NAME)).date()
    start = one_month_before(today)
    return start.strftime("%Y%m%d"), today.strftime("%Y%m%d")


def normalize_date_text(value):
    raw = str(value or "").strip()

    if not raw:
        raise ValueError("日期不能为空。")

    raw = (raw.replace("，", ",").replace("。", ".").replace("／", "/").replace(
        "－", "-").replace("—", "-").replace("年",
                                            "-").replace("月",
                                                         "-").replace("日", ""))

    if re.fullmatch(r"\d{8}", raw):
        d = datetime.strptime(raw, "%Y%m%d").date()
        return d.strftime("%Y%m%d")

    if re.fullmatch(r"\d{6}", raw):
        year = 2000 + int(raw[0:2])
        month = int(raw[2:4])
        day = int(raw[4:6])
        d = datetime(year, month, day).date()
        return d.strftime("%Y%m%d")

    parts = re.findall(r"\d+", raw)
    current_year = datetime.now(ZoneInfo(DEFAULT_TIMEZONE_NAME)).year

    if len(parts) == 3:
        year = int(parts[0])
        month = int(parts[1])
        day = int(parts[2])

        if year < 100:
            year = 2000 + year

        d = datetime(year, month, day).date()
        return d.strftime("%Y%m%d")

    if len(parts) == 2:
        month = int(parts[0])
        day = int(parts[1])
        d = datetime(current_year, month, day).date()
        return d.strftime("%Y%m%d")

    raise ValueError(f"无法识别日期格式：{value}")


def parse_yyyymmdd(value):
    return datetime.strptime(normalize_date_text(value), "%Y%m%d").date()


# ============================================================
# TDX 路径
# ============================================================


def is_tdx_root_dir(path):
    path = normalize_user_path(path, True)
    if not path or not os.path.isdir(path):
        return False

    return os.path.isdir(win_join(path, "vipdoc"))


def resolve_tdx_root_dir(path):
    input_path = normalize_user_path(path, True)

    for folder in iter_parent_dirs(input_path):
        if is_tdx_root_dir(folder):
            return folder

    return input_path


def tdx_vipdoc_dir(tdx_root):
    return win_join(resolve_tdx_root_dir(tdx_root), "vipdoc")


def tdx_hq_cache_dir(tdx_root):
    return win_join(resolve_tdx_root_dir(tdx_root), "T0002", "hq_cache")


def tnf_path(tdx_root, market):
    return win_join(tdx_hq_cache_dir(tdx_root), TNF_FILES[market])


def day_file_path(tdx_root, market, symbol):
    prefix = MARKET_DAY_PREFIX[market]
    return win_join(tdx_vipdoc_dir(tdx_root), prefix, "lday",
                    f"{prefix}{symbol}.day")


# ============================================================
# 配置
# ============================================================


def make_default_config():
    start_date, end_date = default_start_end_dates()

    return {
        "config_version": 1,
        "ui": {
            "window_title": "涨停后新高递归回测工具 v1.0",
            "window_geometry": "1280x760",
        },
        "tdx_root": "",
        "date_range": {
            "start_date": start_date,
            "end_date": end_date,
        },
        "limit_up_threshold": DEFAULT_LIMIT_UP_THRESHOLD,
        "output": {
            "output_dir": "",
            "auto_open_file": DEFAULT_AUTO_OPEN_FILE,
        },
        "path_history": {
            "tdx_root": "",
            "output_dir": "",
        },
    }


def deep_merge_known(default_obj, user_obj):
    result = deepcopy(default_obj)

    if not isinstance(user_obj, dict):
        return result

    for key in result.keys():
        if key not in user_obj:
            continue

        default_value = result[key]
        user_value = user_obj[key]

        if user_value is None:
            continue

        if isinstance(default_value, dict) and isinstance(user_value, dict):
            result[key] = deep_merge_known(default_value, user_value)
        else:
            result[key] = user_value

    return result


def normalize_config_paths(config):
    config["tdx_root"] = normalize_user_path(config.get("tdx_root", ""), True)
    config["output"]["output_dir"] = normalize_user_path(
        config["output"].get("output_dir", ""), True)

    path_history = config.setdefault("path_history", {})
    path_history["tdx_root"] = normalize_user_path(
        path_history.get("tdx_root", ""), True)
    path_history["output_dir"] = normalize_user_path(
        path_history.get("output_dir", ""), True)

    return config


def load_app_config():
    default_config = make_default_config()
    path = get_config_path()

    if not os.path.exists(path):
        return normalize_config_paths(default_config)

    try:
        with open(path, "r", encoding="utf-8") as f:
            saved = json.load(f)
    except Exception:
        return normalize_config_paths(default_config)

    config = deep_merge_known(default_config, saved)

    start_date, end_date = default_start_end_dates()
    config["date_range"]["start_date"] = start_date
    config["date_range"]["end_date"] = end_date

    return normalize_config_paths(config)


def save_app_config(config):
    saved = {
        "config_version":
        1,
        "tdx_root":
        normalize_user_path(config.get("tdx_root", ""), True),
        "limit_up_threshold":
        float(config.get("limit_up_threshold", DEFAULT_LIMIT_UP_THRESHOLD)),
        "output": {
            "output_dir":
            normalize_user_path(config["output"].get("output_dir", ""), True),
            "auto_open_file":
            bool(config["output"].get("auto_open_file", True)),
        },
        "path_history": {
            "tdx_root":
            normalize_user_path(
                config.get("path_history", {}).get("tdx_root", ""), True),
            "output_dir":
            normalize_user_path(
                config.get("path_history", {}).get("output_dir", ""), True),
        },
    }

    with open(get_config_path(), "w", encoding="utf-8") as f:
        json.dump(saved, f, ensure_ascii=False, indent=2)


# ============================================================
# TDX 名称解析
# ============================================================


def decode_ascii(raw):
    try:
        return raw.decode("ascii", errors="ignore").strip("\x00 ").strip()
    except Exception:
        return ""


def decode_gbk(raw):
    try:
        return raw.decode("gbk", errors="ignore").strip("\x00 ").strip()
    except Exception:
        return ""


def load_tnf_name_map_for_market(tdx_root, market):
    path = tnf_path(tdx_root, market)

    if not os.path.exists(path):
        return {}

    with open(path, "rb") as f:
        raw = f.read()

    if len(raw) <= TDX_TNF_HEADER_SIZE:
        return {}

    payload = raw[TDX_TNF_HEADER_SIZE:]
    total = len(payload) // TDX_TNF_RECORD_SIZE
    result = {}

    for idx in range(total):
        start = idx * TDX_TNF_RECORD_SIZE
        rec = payload[start:start + TDX_TNF_RECORD_SIZE]

        if len(rec) < TDX_TNF_RECORD_SIZE:
            continue

        symbol = decode_ascii(rec[0:20])
        name = decode_gbk(rec[31:63])

        if re.fullmatch(r"\d{6}", symbol or ""):
            result[(market, symbol)] = name

    return result


def load_tdx_name_map(tdx_root, log_func):
    result = {}

    for market in ["SH", "SZ"]:
        mp = load_tnf_name_map_for_market(tdx_root, market)
        result.update(mp)
        log_func(f"TDX 名称文件解析完成：{market}，数量：{len(mp)}")

    return result


# ============================================================
# TDX .day 解析
# ============================================================


def unpack_day_record(rec):
    trade_date, open_i, high_i, low_i, close_i, amount_f, volume_i, _ = struct.unpack(
        "<IIIIIfII",
        rec,
    )

    return {
        "DATE": int(trade_date),
        "OPEN": float(open_i) / PRICE_DIVISOR,
        "HIGH": float(high_i) / PRICE_DIVISOR,
        "LOW": float(low_i) / PRICE_DIVISOR,
        "CLOSE": float(close_i) / PRICE_DIVISOR,
        "AMOUNT": float(amount_f),
        "VOLUME": int(volume_i),
    }


def read_day_file_all(file_path):
    if not os.path.exists(file_path):
        return []

    size = os.path.getsize(file_path)
    if size <= 0 or size % TDX_DAY_RECORD_SIZE != 0:
        return []

    rows = []

    with open(file_path, "rb") as f:
        total = size // TDX_DAY_RECORD_SIZE

        for _ in range(total):
            rec = f.read(TDX_DAY_RECORD_SIZE)
            if len(rec) != TDX_DAY_RECORD_SIZE:
                continue

            try:
                row = unpack_day_record(rec)
            except Exception:
                continue

            if row["DATE"] > 0:
                rows.append(row)

    rows.sort(key=lambda x: x["DATE"])
    return rows


def enrich_rows_with_pct(rows):
    """
    给每条日线补充 PCT_CHG。
    第一条无法计算，PCT_CHG 置空。
    """
    result = []

    for i, row in enumerate(rows):
        item = dict(row)

        if i == 0:
            item["PCT_CHG"] = None
        else:
            prev_close = float(rows[i - 1]["CLOSE"])
            if prev_close > 0:
                item["PCT_CHG"] = (float(row["CLOSE"]) -
                                   prev_close) / prev_close * 100.0
            else:
                item["PCT_CHG"] = None

        result.append(item)

    return result


# ============================================================
# 标的范围
# ============================================================


def is_target_star_or_chinext(market, symbol):
    """
    仅科创板、创业板：
    - 科创板：SH688
    - 创业板：SZ300 / SZ301 / SZ302
    """
    symbol = str(symbol or "")

    if market == "SH" and symbol.startswith("688"):
        return True

    if market == "SZ" and (symbol.startswith("300") or symbol.startswith("301")
                           or symbol.startswith("302")):
        return True

    return False


def collect_target_symbols_from_day_files(tdx_root, log_func):
    vipdoc = tdx_vipdoc_dir(tdx_root)

    result = []

    for market in ["SH", "SZ"]:
        prefix = MARKET_DAY_PREFIX[market]
        folder = win_join(vipdoc, prefix, "lday")

        if not os.path.isdir(folder):
            continue

        pattern = win_join(folder, f"{prefix}*.day")
        files = glob.glob(pattern)

        for path in files:
            filename = ntpath.basename(path).lower()
            m = re.fullmatch(rf"{prefix}(\d{{6}})\.day", filename)
            if not m:
                continue

            symbol = m.group(1)

            if not is_target_star_or_chinext(market, symbol):
                continue

            result.append((market, symbol, path))

    result.sort(key=lambda x: (x[0], x[1]))

    log_func(f"扫描到科创板/创业板 day 文件数量：{len(result)}")
    return result


# ============================================================
# 回测核心逻辑
# ============================================================


def is_limit_up_day(row, threshold):
    pct = row.get("PCT_CHG")
    if pct is None:
        return False
    return float(pct) >= float(threshold)


def in_date_range(row, start_int, end_int):
    d = int(row["DATE"])
    return start_int <= d <= end_int

def has_large_gain_in_previous_days(rows_all, current_index, lookback_days, pct_threshold):
    """
    判断 current_index 之前的过去 lookback_days 个交易日内，
    是否出现过 PCT_CHG >= pct_threshold 的涨幅。

    注意：
    - 不包含 current_index 当天；
    - 如果历史不足 lookback_days，则按已有历史检查。
    """
    start = max(0, current_index - int(lookback_days))

    for k in range(start, current_index):
        pct = rows_all[k].get("PCT_CHG")

        if pct is None:
            continue

        if float(pct) >= float(pct_threshold):
            return True

    return False

def backtest_one_symbol(
    market,
    symbol,
    name,
    rows_all,
    start_date,
    end_date,
    limit_up_threshold,
):
    """
    每只股票回测。

    入选特征：
    1. 第一天收盘价涨停，即 PCT_CHG >= limit_up_threshold；
    2. 第二天最高价严格高于第一天收盘价；
    3. 第二天必须是阳线，即第二天收盘价严格高于第二天开盘价；
    4. 第一天之前过去 10 个交易日内，不得出现 18% 以上涨幅。

    新逻辑：
    - 入选特征点仍然按原规则递归检索；
    - 每找到一个入选特征点后，只向右查找一次“再新高”；
    - 新高：入选特征的第二天；
    - 再新高：入选特征第二天之后，第一次出现 HIGH 严格高于“前一新高日收盘价”的交易日；
    - 找到第一次再新高后立即停止该入选点的后续扫描；
    - 如果没有再新高，则再新高间隔记为 0。
    """
    start_int = int(start_date)
    end_int = int(end_date)

    result_rows = []

    n = len(rows_all)
    scan_start = 0
    event_id = 0

    while scan_start < n - 1:
        found_i = None

        # ----------------------------------------------------
        # 从 scan_start 开始查找下一个入选特征点
        # ----------------------------------------------------
        for i in range(scan_start, n - 1):
            row1 = rows_all[i]
            row2 = rows_all[i + 1]

            # 涨停日和第二日都必须在指定日期范围内
            if not in_date_range(row1, start_int, end_int):
                continue

            if not in_date_range(row2, start_int, end_int):
                continue

            # 条件1：第一天涨停
            if not is_limit_up_day(row1, limit_up_threshold):
                continue

            # 排除条件：过去 10 个交易日内出现过 18% 以上涨幅
            if has_large_gain_in_previous_days(
                rows_all=rows_all,
                current_index=i,
                lookback_days=DEFAULT_EXCLUDE_LOOKBACK_DAYS,
                pct_threshold=DEFAULT_EXCLUDE_PCT_THRESHOLD,
            ):
                continue

            # 条件2：第二天最高价必须严格高于第一天收盘价
            if float(row2["HIGH"]) <= float(row1["CLOSE"]):
                continue

            # 条件3：第二天必须是阳线，即收盘价严格高于开盘价
            if float(row2["CLOSE"]) <= float(row2["OPEN"]):
                continue

            found_i = i
            break

        if found_i is None:
            break

        # ----------------------------------------------------
        # 找到一个入选特征点
        # ----------------------------------------------------
        event_id += 1
        i = found_i

        limit_row = rows_all[i]
        high_row = rows_all[i + 1]

        full_symbol = f"{market}{symbol}"
        event_key = f"{full_symbol}_{limit_row['DATE']}_{event_id}"

        # ----------------------------------------------------
        # 涨停日信息
        # ----------------------------------------------------
        limit_date = str(limit_row["DATE"])
        limit_close = float(limit_row["CLOSE"])
        limit_pct = (
            float(limit_row["PCT_CHG"]) / 100.0
            if limit_row.get("PCT_CHG") is not None
            else None
        )
        limit_amount = float(limit_row.get("AMOUNT", 0) or 0)

        # ----------------------------------------------------
        # 新高日，即入选特征的第二天
        # ----------------------------------------------------
        high_date = str(high_row["DATE"])
        high_open = float(high_row["OPEN"])
        high_close = float(high_row["CLOSE"])
        high_day_pct = (
            float(high_row["PCT_CHG"]) / 100.0
            if high_row.get("PCT_CHG") is not None
            else None
        )
        high_price = float(high_row["HIGH"])
        high_relative_pct = (
            (high_price - limit_close) / limit_close
            if limit_close > 0
            else None
        )
        high_amount = float(high_row.get("AMOUNT", 0) or 0)

        # ----------------------------------------------------
        # 再新高：只找第一次
        # ----------------------------------------------------
        rehigh_date = ""
        rehigh_open = ""
        rehigh_close = ""
        rehigh_day_pct = ""
        rehigh_price = ""
        rehigh_relative_pct = ""
        rehigh_amount = ""
        rehigh_interval = 0

        for j in range(i + 2, n):
            row = rows_all[j]
            row_date_int = int(row["DATE"])

            if row_date_int > end_int:
                break

            if row_date_int < start_int:
                continue

            current_open = float(row["OPEN"])
            current_high = float(row["HIGH"])

            # 再新高：
            # 1. 当日最高价严格高于新高日收盘价；
            # 2. 当日开盘价严格低于新高日收盘价。
            if current_high > high_close and current_open < high_close:
                rehigh_date = str(row["DATE"])
                rehigh_open = current_open
                rehigh_close = float(row["CLOSE"])
                rehigh_day_pct = (
                    float(row["PCT_CHG"]) / 100.0
                    if row.get("PCT_CHG") is not None
                    else None
                )
                rehigh_price = current_high
                rehigh_relative_pct = (
                    (current_high - high_close) / high_close
                    if high_close > 0
                    else None
                )
                rehigh_amount = float(row.get("AMOUNT", 0) or 0)
                rehigh_interval = j - (i + 1)

                # 只找第一次再新高，找到后立即停止
                break

        result_rows.append({
            "事件ID": event_key,
            "完整代码": full_symbol,
            "名称": name,

            "涨停日期": limit_date,
            "涨停收盘价": limit_close,
            "涨停日涨幅": limit_pct,
            "涨停日成交额": limit_amount,

            "新高日期": high_date,
            "新高开盘价": high_open,
            "新高收盘价": high_close,
            "新高当日涨幅": high_day_pct,
            "新高最高价": high_price,
            "新高相对涨幅": high_relative_pct,
            "新高成交额": high_amount,

            "再新高日期": rehigh_date,
            "再新高开盘价": rehigh_open,
            "再新高收盘价": rehigh_close,
            "再新高当日涨幅": rehigh_day_pct,
            "再新高最高价": rehigh_price,
            "再新高相对涨幅": rehigh_relative_pct,
            "再新高成交额": rehigh_amount,
            "再新高间隔": rehigh_interval,
        })

        # ----------------------------------------------------
        # 入选特征点递归扫描规则不变：
        # 下一轮从上一轮第二日，即新高日开始
        # ----------------------------------------------------
        scan_start = i + 1

    return result_rows


def run_backtest(config, output_path, log_func=print):
    validate_config(config, output_path)

    tdx_root = resolve_tdx_root_dir(config["tdx_root"])
    start_date = normalize_date_text(config["date_range"]["start_date"])
    end_date = normalize_date_text(config["date_range"]["end_date"])
    limit_up_threshold = float(config["limit_up_threshold"])

    log_func("开始执行回测。")
    log_func(f"TDX 路径：{tdx_root}")
    log_func(f"日期范围：{start_date} ~ {end_date}")
    log_func(f"涨停判断阈值：涨幅 >= {limit_up_threshold}%")
    log_func(
        "入选特征：第一日涨停收盘；第二日最高价严格高于第一日收盘价；"
        "第二日收盘价严格高于开盘价；且第一日前过去10个交易日内未出现18%以上涨幅"
    )
    log_func("后续扫描：每个入选特征点后只查找第一次再新高；再新高条件为后续日最高价高于新高日收盘价，且后续日开盘价低于新高日收盘价。")
    log_func(f"输出文件：{output_path}")

    name_map = load_tdx_name_map(tdx_root, log_func)
    targets = collect_target_symbols_from_day_files(tdx_root, log_func)

    all_rows = []

    total = len(targets)
    valid_count = 0
    event_count = 0

    for idx, (market, symbol, path) in enumerate(targets, start=1):
        rows = read_day_file_all(path)

        if len(rows) < 3:
            if idx % 300 == 0:
                log_func(
                    f"已处理：{idx}/{total}，有效标的：{valid_count}，入选特征点：{event_count}"
                )
            continue

        rows = enrich_rows_with_pct(rows)

        result_rows = backtest_one_symbol(
            market=market,
            symbol=symbol,
            name=name_map.get((market, symbol), ""),
            rows_all=rows,
            start_date=start_date,
            end_date=end_date,
            limit_up_threshold=limit_up_threshold,
        )

        if result_rows:
            valid_count += 1
            event_count += len(result_rows)
            all_rows.extend(result_rows)

        if idx % 300 == 0:
            log_func(
                f"已处理：{idx}/{total}，有效标的：{valid_count}，入选特征点：{event_count}"
            )

    found_count = sum(1 for x in all_rows if int(x.get("再新高间隔") or 0) > 0)
    not_found_count = len(all_rows) - found_count

    log_func(
        f"扫描完成：目标文件 {total} 个，有入选结果标的 {valid_count} 个，入选特征点 {event_count} 个。"
    )
    log_func(f"找到再新高：{found_count} 条")
    log_func(f"未找到再新高：{not_found_count} 条")
    log_func("开始写入 Excel...")

    total_event_count = len(all_rows)

    found_ratio = (
        found_count / total_event_count
        if total_event_count > 0
        else 0
    )

    not_found_ratio = (
        not_found_count / total_event_count
        if total_event_count > 0
        else 0
    )

    export_excel(
        output_path=output_path,
        result_rows=all_rows,
        params={
            "TDX路径": tdx_root,
            "开始日期": start_date,
            "结束日期": end_date,
            "涨停判断阈值": f"涨幅 >= {limit_up_threshold}%",
            "入选特征": "第一日涨停收盘；第二日最高价严格高于第一日收盘价；第二日收盘价严格高于开盘价；第一日前过去10个交易日内未出现18%以上涨幅",
            "后续扫描规则": "每个入选特征点后，只查找第一次再新高；再新高条件为后续日最高价高于新高日收盘价，且后续日开盘价低于新高日收盘价；找到一次即停止；未找到则再新高间隔为0",
            "目标范围": "科创板 SH688；创业板 SZ300/SZ301/SZ302",

            "目标文件数": total,
            "有入选结果标的数": valid_count,
            "入选特征点总数": total_event_count,
            "找到再新高数量": found_count,
            "未找到再新高数量": not_found_count,
            "找到再新高占比": f"{found_ratio:.2%}",
            "未找到再新高占比": f"{not_found_ratio:.2%}",

            "生成时间": datetime.now(
                ZoneInfo(DEFAULT_TIMEZONE_NAME)
            ).strftime("%Y-%m-%d %H:%M:%S"),
        },
    )

    log_func("-" * 60)
    log_func("导出完成。")
    log_func(f"入选特征点总数：{len(all_rows)}")
    log_func(f"找到再新高：{found_count}")
    log_func(f"未找到再新高：{not_found_count}")
    log_func(f"文件位置：{output_path}")


# ============================================================
# Excel 输出
# ============================================================


def excel_display_text(value, number_format=None):
    if value is None:
        return ""

    if isinstance(value, float) or isinstance(value, int):
        if number_format == "0.00%":
            return f"{float(value) * 100:.2f}%"

        if number_format == "0.00":
            return f"{float(value):.2f}"

        if number_format == "#,##0.00":
            return f"{float(value):,.2f}"

        if number_format == "#,##0":
            return f"{int(value):,}"

    return str(value)


def text_visual_width(text):
    text = str(text or "")
    width = 0

    for ch in text:
        if ord(ch) > 127:
            width += 2
        else:
            width += 1

    return width


def style_common_sheet(
    ws,
    percent_columns=None,
    price_columns=None,
    int_columns=None,
    amount_columns=None,
):
    percent_columns = set(percent_columns or [])
    price_columns = set(price_columns or [])
    int_columns = set(int_columns or [])
    amount_columns = set(amount_columns or [])

    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    # 字段分组底色
    limit_fill = PatternFill("solid", fgColor="FFF2CC")   # 涨停日字段：浅黄
    high_fill = PatternFill("solid", fgColor="E2F0D9")    # 新高字段：浅绿
    rehigh_fill = PatternFill("solid", fgColor="DDEBF7")  # 再新高字段：浅蓝

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_map = {cell.value: cell.column for cell in ws[1]}

    for title in percent_columns:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col)
                if cell.value != "":
                    cell.number_format = "0.00%"

    for title in price_columns:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col)
                if cell.value != "":
                    cell.number_format = "0.00"

    for title in int_columns:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col)
                if cell.value != "":
                    cell.number_format = "#,##0"

    for title in amount_columns:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col)
                if cell.value != "":
                    cell.number_format = "#,##0.00"

    # 字段列分组
    limit_cols = set()
    high_cols = set()
    rehigh_cols = set()

    for title, col in header_map.items():
        title_text = str(title or "")

        if title_text.startswith("涨停"):
            limit_cols.add(col)
        elif title_text.startswith("新高"):
            high_cols.add(col)
        elif title_text.startswith("再新高"):
            rehigh_cols.add(col)

    for row_idx in range(1, ws.max_row + 1):
        is_header = row_idx == 1

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border

            if is_header:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                continue

            # 数据区按字段组涂色
            if col_idx in limit_cols:
                cell.fill = limit_fill
            elif col_idx in high_cols:
                cell.fill = high_fill
            elif col_idx in rehigh_cols:
                cell.fill = rehigh_fill

            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header_value = ws.cell(1, col_idx).value
        max_width = text_visual_width(header_value)

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            display_text = excel_display_text(cell.value, cell.number_format)
            max_width = max(max_width, text_visual_width(display_text))

        width = max_width + 2
        width = max(8, min(width, 36))
        ws.column_dimensions[letter].width = width


def export_excel(output_path, result_rows, params):
    columns = [
        "事件ID",
        "完整代码",
        "名称",

        "涨停日期",
        "涨停收盘价",
        "涨停日涨幅",
        "涨停日成交额",

        "新高日期",
        "新高开盘价",
        "新高收盘价",
        "新高当日涨幅",
        "新高最高价",
        "新高相对涨幅",
        "新高成交额",

        "再新高日期",
        "再新高开盘价",
        "再新高收盘价",
        "再新高当日涨幅",
        "再新高最高价",
        "再新高相对涨幅",
        "再新高成交额",
        "再新高间隔",
    ]

    param_rows = [{"参数": k, "值": v} for k, v in params.items()]

    if result_rows:
        df_all = pd.DataFrame(result_rows, columns=columns)
    else:
        df_all = pd.DataFrame(columns=columns)

    if not df_all.empty:
        df_found = df_all[df_all["再新高间隔"].fillna(0).astype(int) > 0].copy()
        df_not_found = df_all[df_all["再新高间隔"].fillna(0).astype(int) <= 0].copy()

        # 找到再新高表格排序：
        # 第一条件：再新高间隔升序
        # 第二条件：新高日期降序
        # 第三条件：完整代码升序
        df_found.sort_values(
            by=["再新高间隔", "新高日期", "完整代码"],
            ascending=[True, False, True],
            inplace=True,
        )

        # 未找到再新高表格排序：
        # 第一条件：按新高日期降序
        # 第二条件：完整代码升序
        df_not_found.sort_values(
            by=["新高日期", "完整代码"],
            ascending=[False, True],
            inplace=True,
        )

    else:
        df_found = pd.DataFrame(columns=columns)
        df_not_found = pd.DataFrame(columns=columns)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_found.to_excel(writer, sheet_name="找到再新高", index=False)
        df_not_found.to_excel(writer, sheet_name="未找到再新高", index=False)
        pd.DataFrame(param_rows).to_excel(
            writer,
            sheet_name="参数及结果",
            index=False,
        )

        common_percent_columns = [
            "涨停日涨幅",
            "新高当日涨幅",
            "新高相对涨幅",
            "再新高当日涨幅",
            "再新高相对涨幅",
        ]

        common_price_columns = [
            "涨停收盘价",
            "新高开盘价",
            "新高收盘价",
            "新高最高价",
            "再新高开盘价",
            "再新高收盘价",
            "再新高最高价",
        ]

        common_amount_columns = [
            "涨停日成交额",
            "新高成交额",
            "再新高成交额",
        ]

        style_common_sheet(
            writer.sheets["找到再新高"],
            percent_columns=common_percent_columns,
            price_columns=common_price_columns,
            int_columns=[
                "再新高间隔",
            ],
            amount_columns=common_amount_columns,
        )

        style_common_sheet(
            writer.sheets["未找到再新高"],
            percent_columns=common_percent_columns,
            price_columns=common_price_columns,
            int_columns=[
                "再新高间隔",
            ],
            amount_columns=common_amount_columns,
        )

        style_common_sheet(writer.sheets["参数及结果"])

        writer.book.active = writer.book.sheetnames.index("找到再新高")


# ============================================================
# 校验与输出路径
# ============================================================


def validate_tdx_daily_source(tdx_root):
    tdx_root = resolve_tdx_root_dir(tdx_root)

    if not tdx_root:
        raise ValueError("TDX 路径不能为空。")

    if not os.path.isdir(tdx_root):
        raise FileNotFoundError(f"TDX 目录不存在：\n{tdx_root}")

    vipdoc = tdx_vipdoc_dir(tdx_root)
    if not os.path.isdir(vipdoc):
        raise FileNotFoundError(f"TDX vipdoc 目录不存在：\n{vipdoc}")

    sh_lday = win_join(vipdoc, "sh", "lday")
    sz_lday = win_join(vipdoc, "sz", "lday")

    if not os.path.isdir(sh_lday):
        raise FileNotFoundError(f"TDX 沪市日线目录不存在：\n{sh_lday}")

    if not os.path.isdir(sz_lday):
        raise FileNotFoundError(f"TDX 深市日线目录不存在：\n{sz_lday}")

    return tdx_root


def validate_config(config, output_path):
    tdx_root = validate_tdx_daily_source(config["tdx_root"])
    config["tdx_root"] = tdx_root

    start_date = parse_yyyymmdd(config["date_range"]["start_date"])
    end_date = parse_yyyymmdd(config["date_range"]["end_date"])

    if start_date > end_date:
        raise ValueError("开始日期不能晚于结束日期。")

    threshold = float(config["limit_up_threshold"])
    if threshold < -100 or threshold > 100:
        raise ValueError("涨停判断阈值不合理。")

    output_path = normalize_user_path(output_path, True)

    if not output_path:
        raise ValueError("输出文件路径不能为空。")

    if not output_path.lower().endswith(".xlsx"):
        raise ValueError("输出文件必须是 .xlsx 格式。")

    output_dir = win_dirname(output_path)
    if output_dir and not os.path.isdir(output_dir):
        raise FileNotFoundError(f"输出目录不存在：\n{output_dir}")


def default_output_path(config):
    start_str = config["date_range"]["start_date"]
    end_str = config["date_range"]["end_date"]
    filename = f"{start_str}-{end_str}_涨停后新高递归回测.xlsx"

    output_dir = normalize_user_path(config["output"]["output_dir"], True)
    if not output_dir:
        output_dir = get_program_dir()

    return win_join(output_dir, filename)


def make_unique_output_path(path):
    path = normalize_user_path(path, True)
    folder = win_dirname(path)
    filename = ntpath.basename(path)
    stem, ext = ntpath.splitext(filename)

    if not ext:
        ext = ".xlsx"

    idx = 1
    while True:
        candidate = win_join(folder, f"{stem}_{idx}{ext}")
        if not os.path.exists(candidate):
            return candidate
        idx += 1


# ============================================================
# GUI
# ============================================================


class App:

    def __init__(self, root):
        self.root = root
        self.config = load_app_config()
        self.log_queue = queue.Queue()
        self.auto_save_job = None
        self.is_building_ui = True

        self.root.title(self.config["ui"]["window_title"])
        self.set_window_icon()
        self.root.geometry(self.config["ui"]["window_geometry"])

        try:
            self.root.state("zoomed")
        except Exception:
            pass

        self.build_vars()
        self.build_ui()
        self.update_output_path_by_dates()
        self.bind_auto_save_events()
        self.is_building_ui = False

        self.root.after(100, self.poll_log_queue)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def build_vars(self):
        self.tdx_root_var = tk.StringVar(value=self.config["tdx_root"])

        self.start_date_var = tk.StringVar(
            value=self.config["date_range"]["start_date"])
        self.end_date_var = tk.StringVar(
            value=self.config["date_range"]["end_date"])

        self.limit_up_threshold_var = tk.StringVar(value=str(
            self.config.get("limit_up_threshold", DEFAULT_LIMIT_UP_THRESHOLD)))

        self.output_path_var = tk.StringVar(
            value=default_output_path(self.config))
        self.auto_open_file_var = tk.BooleanVar(
            value=self.config["output"].get("auto_open_file", True))

    def build_ui(self):
        main = tk.Frame(self.root)
        main.pack(fill="both", expand=True, padx=10, pady=6)

        title_frame = tk.Frame(main)
        title_frame.pack(fill="x", pady=(0, 4))

        tk.Label(
            title_frame,
            text="涨停后新高递归回测工具——科创板 / 创业板 / TDX day 文件",
            font=("Microsoft YaHei", 15, "bold"),
        ).pack(side="left", anchor="w")

        self.build_data_source_frame(main)
        self.build_param_frame(main)
        self.build_output_frame(main)
        self.build_log_frame(main)

    def build_data_source_frame(self, parent):
        frame = tk.LabelFrame(parent, text="数据源")
        frame.pack(fill="x", pady=3)

        tk.Label(frame, text="TDX 根目录：").grid(row=0,
                                              column=0,
                                              sticky="e",
                                              padx=8,
                                              pady=6)

        tk.Entry(frame, textvariable=self.tdx_root_var,
                 width=120).grid(row=0, column=1, sticky="we", padx=6, pady=6)

        tk.Button(
            frame,
            text="选择文件夹",
            width=UI_BUTTON_WIDTH,
            command=self.choose_tdx_root,
        ).grid(row=0, column=2, sticky="w", padx=4, pady=6)

        tk.Button(
            frame,
            text="打开文件夹",
            width=UI_BUTTON_WIDTH,
            command=lambda: self.open_existing_folder(self.tdx_root_var.get()),
        ).grid(row=0, column=3, sticky="w", padx=4, pady=6)

        frame.columnconfigure(1, weight=1)

    def build_param_frame(self, parent):
        frame = tk.LabelFrame(parent, text="日期与参数")
        frame.pack(fill="x", pady=3)

        tk.Label(frame, text="开始：").grid(row=0,
                                         column=0,
                                         sticky="e",
                                         padx=(6, 2),
                                         pady=6)
        start_entry = tk.Entry(frame,
                               textvariable=self.start_date_var,
                               width=12)
        start_entry.grid(row=0, column=1, sticky="w", padx=(2, 8), pady=6)
        start_entry.bind(
            "<FocusOut>",
            lambda e: self.normalize_date_entry(self.start_date_var))

        tk.Label(frame, text="结束：").grid(row=0,
                                         column=2,
                                         sticky="e",
                                         padx=(6, 2),
                                         pady=6)
        end_entry = tk.Entry(frame, textvariable=self.end_date_var, width=12)
        end_entry.grid(row=0, column=3, sticky="w", padx=(2, 8), pady=6)
        end_entry.bind("<FocusOut>",
                       lambda e: self.normalize_date_entry(self.end_date_var))

        tk.Label(frame, text="涨停判断：涨幅 >=").grid(row=0,
                                                column=4,
                                                sticky="e",
                                                padx=(6, 2),
                                                pady=6)
        tk.Entry(frame, textvariable=self.limit_up_threshold_var,
                 width=8).grid(row=0,
                               column=5,
                               sticky="w",
                               padx=(2, 1),
                               pady=6)
        tk.Label(frame, text="%").grid(row=0,
                                       column=6,
                                       sticky="w",
                                       padx=(0, 8),
                                       pady=6)

        tk.Label(
            frame,
            text="说明：TDX day 无涨停价字段，程序用当日涨幅阈值近似判断涨停，默认 19.8%。",
            fg="#666666",
        ).grid(row=0, column=7, sticky="w", padx=8, pady=6)

        self.start_date_var.trace_add(
            "write", lambda *_: self.update_output_path_by_dates())
        self.end_date_var.trace_add(
            "write", lambda *_: self.update_output_path_by_dates())

    def build_output_frame(self, parent):
        frame = tk.LabelFrame(parent, text="输出")
        frame.pack(fill="x", pady=3)

        tk.Label(frame, text="输出文件：").grid(row=0,
                                           column=0,
                                           sticky="e",
                                           padx=8,
                                           pady=6)

        tk.Entry(frame, textvariable=self.output_path_var,
                 width=120).grid(row=0, column=1, sticky="we", padx=6, pady=6)

        tk.Button(
            frame,
            text="选择文件夹",
            width=UI_BUTTON_WIDTH,
            command=self.choose_output_dir,
        ).grid(row=0, column=2, sticky="w", padx=4, pady=6)

        tk.Button(
            frame,
            text="另存为",
            width=UI_BUTTON_WIDTH,
            command=self.choose_output_file,
        ).grid(row=0, column=3, sticky="w", padx=4, pady=6)

        tk.Button(
            frame,
            text="打开文件夹",
            width=UI_BUTTON_WIDTH,
            command=self.open_output_folder,
        ).grid(row=0, column=4, sticky="w", padx=4, pady=6)

        frame.columnconfigure(1, weight=1)

    def build_log_frame(self, parent):
        frame = tk.LabelFrame(parent, text="日志")
        frame.pack(fill="both", expand=True, pady=3)

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        text_container = tk.Frame(frame)
        text_container.grid(row=0,
                            column=0,
                            sticky="nsew",
                            padx=6,
                            pady=(3, 2))
        text_container.rowconfigure(0, weight=1)
        text_container.columnconfigure(0, weight=1)

        self.log_text = tk.Text(text_container, height=18, wrap="none")
        self.log_text.grid(row=0, column=0, sticky="nsew")

        y_scroll = tk.Scrollbar(text_container,
                                orient="vertical",
                                command=self.log_text.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")

        x_scroll = tk.Scrollbar(text_container,
                                orient="horizontal",
                                command=self.log_text.xview)
        x_scroll.grid(row=1, column=0, sticky="we")

        self.log_text.configure(yscrollcommand=y_scroll.set,
                                xscrollcommand=x_scroll.set)

        btn_frame = tk.Frame(frame)
        btn_frame.grid(row=1, column=0, sticky="e", padx=4, pady=(1, 3))

        tk.Checkbutton(
            btn_frame,
            text="自动打开成果文件",
            variable=self.auto_open_file_var,
        ).pack(side="left", padx=8)

        self.export_btn = tk.Button(
            btn_frame,
            text="开始回测",
            width=UI_MAIN_BUTTON_WIDTH,
            command=self.start_export,
        )
        self.export_btn.pack(side="left", padx=4)

    def set_window_icon(self):
        icon_path = get_resource_path(ICON_FILE_NAME)

        if not os.path.exists(icon_path):
            return

        try:
            self.root.iconbitmap(icon_path)
        except Exception:
            pass

    def get_dialog_initial_dir(self, current_path, history_key):
        current_path = normalize_user_path(current_path, True)

        if current_path:
            if os.path.isfile(current_path):
                current_dir = win_dirname(current_path)
            else:
                current_dir = current_path

            if os.path.isdir(current_dir):
                return current_dir

        history = self.config.get("path_history", {}).get(history_key, "")
        history = normalize_user_path(history, True)

        if history and os.path.isdir(history):
            return history

        return get_program_dir()

    def update_path_history(self, history_key, path):
        path = normalize_user_path(path, True)

        if not path:
            return

        if os.path.isfile(path):
            path = win_dirname(path)

        if not os.path.isdir(path):
            return

        self.config.setdefault("path_history", {})[history_key] = path

    def choose_tdx_root(self):
        initial_dir = self.get_dialog_initial_dir(
            current_path=self.tdx_root_var.get(),
            history_key="tdx_root",
        )

        path = filedialog.askdirectory(
            title="选择 TDX 根目录或 vipdoc 下级目录",
            initialdir=initial_dir,
        )

        if path:
            resolved = resolve_tdx_root_dir(path)
            self.tdx_root_var.set(resolved)
            self.update_path_history("tdx_root", resolved)
            self.schedule_auto_save()

    def choose_output_dir(self):
        initial_dir = self.get_dialog_initial_dir(
            current_path=win_dirname(self.output_path_var.get()),
            history_key="output_dir",
        )

        path = filedialog.askdirectory(
            title="选择输出目录",
            initialdir=initial_dir,
        )

        if path:
            path = normalize_user_path(path, True)
            self.config["output"]["output_dir"] = path
            self.update_path_history("output_dir", path)
            self.update_output_path_by_dates(force_dir=path)
            self.schedule_auto_save()

    def choose_output_file(self):
        try:
            start_str = normalize_date_text(self.start_date_var.get())
            end_str = normalize_date_text(self.end_date_var.get())
        except Exception:
            start_str = self.start_date_var.get().strip()
            end_str = self.end_date_var.get().strip()

        initial_dir = self.get_dialog_initial_dir(
            current_path=self.output_path_var.get(),
            history_key="output_dir",
        )

        path = filedialog.asksaveasfilename(
            title="选择输出 Excel 文件",
            defaultextension=".xlsx",
            initialdir=initial_dir,
            initialfile=f"{start_str}-{end_str}_涨停后新高递归回测.xlsx",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("所有文件", "*.*"),
            ],
        )

        if path:
            path = normalize_user_path(path, True)
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.output_path_var.set(path)

            output_dir = normalize_user_path(win_dirname(path), True)
            self.config["output"]["output_dir"] = output_dir
            self.update_path_history("output_dir", output_dir)
            self.schedule_auto_save()

    def open_output_folder(self):
        output_path = normalize_user_path(self.output_path_var.get(), True)
        folder = win_dirname(
            output_path
        ) if output_path else self.config["output"]["output_dir"]
        self.open_existing_folder(folder)

    def open_existing_folder(self, folder):
        folder = normalize_user_path(folder, True)

        if not folder:
            messagebox.showwarning("无法打开", "路径为空。")
            return

        if not os.path.isdir(folder):
            messagebox.showwarning("无法打开", f"文件夹不存在：\n{folder}")
            return

        try:
            os.startfile(folder)
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件夹：\n{folder}\n\n原因：{e}")

    def normalize_date_entry(self, var):
        raw = var.get().strip()
        if not raw:
            return

        try:
            var.set(normalize_date_text(raw))
            self.update_output_path_by_dates()
        except Exception:
            pass

    def update_output_path_by_dates(self, force_dir=None):
        try:
            start_str = normalize_date_text(self.start_date_var.get())
            end_str = normalize_date_text(self.end_date_var.get())
        except Exception:
            return

        output_dir = force_dir if force_dir is not None else self.config[
            "output"]["output_dir"]
        output_dir = normalize_user_path(output_dir, True)

        if not output_dir:
            output_dir = get_program_dir()

        self.output_path_var.set(
            win_join(output_dir, f"{start_str}-{end_str}_涨停后新高递归回测.xlsx"))

    def collect_config_from_ui(self):
        start = normalize_date_text(self.start_date_var.get())
        end = normalize_date_text(self.end_date_var.get())

        if parse_yyyymmdd(start) > parse_yyyymmdd(end):
            raise ValueError("开始日期不能晚于结束日期。")

        try:
            threshold = float(self.limit_up_threshold_var.get().strip())
        except Exception:
            raise ValueError("涨停判断阈值必须是数字。")

        config = deepcopy(self.config)

        config["tdx_root"] = resolve_tdx_root_dir(self.tdx_root_var.get())
        config["date_range"]["start_date"] = start
        config["date_range"]["end_date"] = end
        config["limit_up_threshold"] = threshold

        output_path = normalize_user_path(self.output_path_var.get(), True)
        if output_path and not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"
            self.output_path_var.set(output_path)

        config["output"]["output_dir"] = normalize_user_path(
            win_dirname(output_path), True)
        config["output"]["auto_open_file"] = bool(
            self.auto_open_file_var.get())

        return normalize_config_paths(config)

    def resolve_existing_output_file(self, output_path):
        output_path = normalize_user_path(output_path, True)

        if not os.path.exists(output_path):
            return output_path

        choice = messagebox.askyesnocancel(
            "文件已存在",
            "输出文件已存在：\n\n"
            f"{output_path}\n\n"
            "请选择处理方式：\n\n"
            "是：覆盖原文件\n"
            "否：自动改名保存\n"
            "取消：取消本次导出",
        )

        if choice is True:
            return output_path

        if choice is False:
            new_path = make_unique_output_path(output_path)
            self.output_path_var.set(new_path)
            return new_path

        return None

    def can_write_output_file(self, output_path):
        output_path = normalize_user_path(output_path, True)

        try:
            folder = win_dirname(output_path)
            if folder and not os.path.isdir(folder):
                return False

            if os.path.exists(output_path):
                with open(output_path, "a+b"):
                    pass
            else:
                with open(output_path, "wb"):
                    pass
                os.remove(output_path)

            return True
        except Exception:
            return False

    def resolve_writable_output_file(self, output_path):
        current_path = normalize_user_path(output_path, True)

        while True:
            if self.can_write_output_file(current_path):
                return current_path

            choice = messagebox.askyesnocancel(
                "文件无法写入",
                "当前输出文件无法写入，可能正在被 Excel 或其他程序打开：\n\n"
                f"{current_path}\n\n"
                "请选择处理方式：\n\n"
                "已经关闭文件选【是】\n"
                "要改名另存选【否】\n"
                "中止本次导出选【取消】",
            )

            if choice is True:
                continue

            if choice is False:
                current_path = make_unique_output_path(current_path)
                self.output_path_var.set(current_path)
                continue

            return None

    def bind_auto_save_events(self):
        vars_to_watch = [
            self.tdx_root_var,
            self.start_date_var,
            self.end_date_var,
            self.limit_up_threshold_var,
            self.output_path_var,
            self.auto_open_file_var,
        ]

        for var in vars_to_watch:
            var.trace_add("write", lambda *_: self.schedule_auto_save())

    def schedule_auto_save(self):
        if getattr(self, "is_building_ui", False):
            return

        if self.auto_save_job is not None:
            self.root.after_cancel(self.auto_save_job)

        self.auto_save_job = self.root.after(500,
                                             self.auto_save_current_config)

    def auto_save_current_config(self):
        self.auto_save_job = None

        try:
            self.config = self.collect_config_from_ui()
            save_app_config(self.config)
        except Exception:
            pass

    def log(self, msg):
        self.log_queue.put(str(msg))

    def poll_log_queue(self):
        while True:
            try:
                msg = self.log_queue.get_nowait()
            except queue.Empty:
                break

            self.log_text.insert("end", msg + "\n")
            self.log_text.see("end")

        self.root.after(100, self.poll_log_queue)

    def start_export(self):
        try:
            config = self.collect_config_from_ui()
            output_path = normalize_user_path(self.output_path_var.get(), True)

            if not output_path.lower().endswith(".xlsx"):
                output_path += ".xlsx"
                self.output_path_var.set(output_path)

            validate_config(config, output_path)

            resolved = self.resolve_existing_output_file(output_path)
            if resolved is None:
                return

            writable_path = self.resolve_writable_output_file(resolved)
            if writable_path is None:
                return

            output_path = normalize_user_path(writable_path, True)
            self.output_path_var.set(output_path)

            config["output"]["output_dir"] = normalize_user_path(
                win_dirname(output_path), True)

            self.config = config
            save_app_config(self.config)

        except Exception as e:
            messagebox.showerror("输入错误", str(e))
            return

        self.export_btn.config(state="disabled")
        self.log_text.delete("1.0", "end")

        thread = threading.Thread(
            target=self.run_export_thread,
            args=(deepcopy(self.config), output_path),
            daemon=True,
        )
        thread.start()

    def run_export_thread(self, config, output_path):
        try:
            run_backtest(config, output_path, log_func=self.log)
            self.log("任务成功完成。")

            auto_open = bool(config["output"].get("auto_open_file", True))

            def after_success():
                if auto_open:
                    try:
                        os.startfile(output_path)
                    except Exception as e:
                        self.log(f"文件已生成，但自动打开失败：{e}")
                        messagebox.showinfo("完成", "导出完成，但自动打开文件失败。")
                        return

                    messagebox.showinfo("完成", "导出完成，已自动打开文件。")
                else:
                    messagebox.showinfo("完成", "导出完成。")

            self.root.after(0, after_success)

        except Exception as e:
            msg = str(e)
            self.log(f"导出失败：{msg}")
            self.root.after(0, lambda m=msg: messagebox.showerror("导出失败", m))

        finally:
            self.root.after(0, lambda: self.export_btn.config(state="normal"))

    def on_close(self):
        try:
            self.config = self.collect_config_from_ui()
            save_app_config(self.config)
        except Exception:
            pass

        self.root.destroy()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()

# 以下内容可按你的项目路径自行修改保留
# 【运行虚拟环境】
# PS E:\AppProject\LimitUpNewHighBacktest> .venv\Scripts\activate
#
# 【封装 exe 文件示例】
# Set-Location "E:\AppProject\LimitUpNewHighBacktest"
# pyinstaller "E:\AppProject\LimitUpNewHighBacktest\LimitUpNewHighBacktest.py" --onefile --windowed --clean --noconfirm --name "LimitUpNewHighBacktest" --icon "E:\AppProject\LimitUpNewHighBacktest\icon.ico" --add-data "E:\AppProject\LimitUpNewHighBacktest\icon.ico;." --hidden-import openpyxl.styles --exclude-module matplotlib --exclude-module scipy --exclude-module PyQt5 --exclude-module PyQt6 --exclude-module PySide2 --exclude-module PySide6 --exclude-module IPython --exclude-module notebook --exclude-module pytest --exclude-module unittest --exclude-module pydoc --exclude-module doctest --exclude-module html --exclude-module http --exclude-module xmlrpc
