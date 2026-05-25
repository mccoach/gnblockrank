import os
import re
import sys
import json
import glob
import struct
import queue
import ntpath
import threading
import calendar
from copy import deepcopy
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 常量
# ============================================================

CONFIG_FILE_NAME = "GNBlockRank_config.json"
QR_CODE_FILE_NAME = "wechat_qr.png"
ICON_FILE_NAME = "icon.ico"

DEFAULT_TIMEZONE_NAME = "Asia/Shanghai"

DAILY_SOURCE_TDX = "TDX"

CONCEPT_SOURCE_TDX = "TDX"
CONCEPT_SOURCE_THS = "THS"

TDX_DAY_RECORD_SIZE = 32
TDX_TNF_HEADER_SIZE = 50
TDX_TNF_RECORD_SIZE = 360

PRICE_DIVISOR = 100.0

DEFAULT_ENTRY_THRESHOLD = 9.7
DEFAULT_MARK_THRESHOLD = 19.98
DEFAULT_MARK_FILL_COLOR = "FF8080"

UI_BUTTON_WIDTH = 10
UI_MAIN_BUTTON_WIDTH = 14

SUMMARY_SHEET_NAMES = ["近5日汇总", "近10日汇总", "全部汇总"]
SUMMARY_MAX_RANK = 50

CONCEPT_BLOCK_NAME_BLACKLIST = {
    "融资融券",
    "深股通",
    "沪股通",
}

MARKET_ID_MAP = {
    "0": "SZ",
    "1": "SH",
    "2": "BJ",
}

MARKET_DAY_PREFIX = {
    "SH": "sh",
    "SZ": "sz",
    "BJ": "bj",
}

TNF_FILES = {
    "SH": "shs.tnf",
    "SZ": "szs.tnf",
    "BJ": "bjs.tnf",
}

THS_CONCEPT_BLOCK_FILE = "block_2B.ini"

THS_A_STOCK_PREFIX_TO_MARKET = {
    "17": "SH",
    "22": "SH",
    "33": "SZ",
    "105": "BJ",
}

THS_BLOCK_PREFIX_TO_STOCKNAME_MARKET_ID = {
    "17": "16",
    "22": "16",
    "33": "32",
    "105": "144",
}

THS_NAME_SECTION_TO_MARKET = {
    ("16", "17"): "SH",
    ("16", "22"): "SH",
    ("32", "33"): "SZ",
    ("144", "151"): "BJ",
}

BLOCK_FIELD_DEFS = [
    {
        "key": "RANK",
        "title": "序号"
    },
    {
        "key": "BLOCK_NAME",
        "title": "概念板块名称"
    },
    {
        "key": "MEMBER_COUNT",
        "title": "总成分股数"
    },
    {
        "key": "QUALIFIED_COUNT",
        "title": "达标个股数"
    },
    {
        "key": "QUALIFIED_RATIO",
        "title": "达标占比"
    },
]

DETAIL_FIELD_DEFS = [
    {
        "key": "SYMBOL_FULL",
        "title": "标的代码"
    },
    {
        "key": "NAME",
        "title": "标的名称"
    },
    {
        "key": "PCT_CHG",
        "title": "当日涨幅"
    },
    {
        "key": "OPEN",
        "title": "开盘价"
    },
    {
        "key": "HIGH",
        "title": "最高价"
    },
    {
        "key": "LOW",
        "title": "最低价"
    },
    {
        "key": "CLOSE",
        "title": "收盘价"
    },
    {
        "key": "VOLUME",
        "title": "成交量"
    },
    {
        "key": "AMOUNT",
        "title": "成交额"
    },
    {
        "key": "APPEAR_COUNT",
        "title": "出现次数"
    },
]

# ============================================================
# 路径工具
# ============================================================


def normalize_path_slashes(path):
    if path is None:
        return ""
    return str(path).replace("/", "\\")


def strip_outer_path_quotes(path):
    text = str(path or "").strip()
    pairs = [
        ('"', '"'),
        ("'", "'"),
        ("“", "”"),
        ("‘", "’"),
    ]

    changed = True
    while changed and len(text) >= 2:
        changed = False
        for left, right in pairs:
            if text.startswith(left) and text.endswith(right):
                text = text[1:-1].strip()
                changed = True
                break

    return text


def is_windows_root_path(path):
    path = normalize_path_slashes(path).strip()
    if not path:
        return False

    drive, tail = ntpath.splitdrive(path)
    if drive and tail == "\\":
        return True

    if path.startswith("\\\\"):
        parts = [p for p in path.split("\\") if p]
        return len(parts) == 2

    return False


def strip_redundant_trailing_slashes(path):
    text = normalize_path_slashes(path).strip()

    while len(text) > 1 and text.endswith(
            "\\") and not is_windows_root_path(text):
        text = text[:-1]

    return text


def normalize_user_path(path, trim_trailing_slash=True):
    text = strip_outer_path_quotes(path)
    text = normalize_path_slashes(text).strip()

    if trim_trailing_slash:
        text = strip_redundant_trailing_slashes(text)

    return text


def win_join(*parts):
    return normalize_path_slashes(ntpath.join(*parts))


def win_dirname(path):
    return normalize_path_slashes(ntpath.dirname(path))


def get_program_dir():
    if getattr(sys, "frozen", False):
        return normalize_path_slashes(os.path.dirname(sys.executable))
    return normalize_path_slashes(os.path.dirname(os.path.abspath(__file__)))


def get_resource_path(filename):
    """
    获取资源文件路径。
    普通运行：从 .py 所在目录读取
    PyInstaller onefile：从临时解包目录读取
    """
    if getattr(sys, "frozen", False):
        base_dir = getattr(sys, "_MEIPASS", get_program_dir())
    else:
        base_dir = get_program_dir()

    return normalize_path_slashes(os.path.join(base_dir, filename))


def get_config_path():
    return win_join(get_program_dir(), CONFIG_FILE_NAME)


def existing_path_for_root_search(path):
    path = normalize_user_path(path, True)
    if not path:
        return ""

    if os.path.isfile(path):
        return win_dirname(path)

    return path


def iter_parent_dirs(path):
    current = existing_path_for_root_search(path)

    while current:
        yield current

        parent = win_dirname(current)
        if not parent or parent == current:
            break

        current = parent


# ============================================================
# 日期工具
# ============================================================


def one_month_before(d):
    year = d.year
    month = d.month - 1

    if month == 0:
        year -= 1
        month = 12

    day = min(d.day, calendar.monthrange(year, month)[1])
    return d.replace(year=year, month=month, day=day)


def default_start_end_dates():
    today = datetime.now(ZoneInfo(DEFAULT_TIMEZONE_NAME)).date()
    start = one_month_before(today)
    return start.strftime("%Y%m%d"), today.strftime("%Y%m%d")


def normalize_date_text(value):
    raw = str(value or "").strip()

    if not raw:
        raise ValueError("日期不能为空。")

    raw = (raw.replace("，", ",").replace("。", ".").replace("／", "/").replace(
        "－", "-").replace("—", "-").replace("年",
                                            "-").replace("月",
                                                         "-").replace("日", ""))

    if re.fullmatch(r"\d{8}", raw):
        d = datetime.strptime(raw, "%Y%m%d").date()
        return d.strftime("%Y%m%d")

    if re.fullmatch(r"\d{6}", raw):
        year = 2000 + int(raw[0:2])
        month = int(raw[2:4])
        day = int(raw[4:6])
        d = datetime(year, month, day).date()
        return d.strftime("%Y%m%d")

    parts = re.findall(r"\d+", raw)
    current_year = datetime.now(ZoneInfo(DEFAULT_TIMEZONE_NAME)).year

    if len(parts) == 3:
        year = int(parts[0])
        month = int(parts[1])
        day = int(parts[2])

        if year < 100:
            year = 2000 + year

        d = datetime(year, month, day).date()
        return d.strftime("%Y%m%d")

    if len(parts) == 2:
        month = int(parts[0])
        day = int(parts[1])
        d = datetime(current_year, month, day).date()
        return d.strftime("%Y%m%d")

    raise ValueError(f"无法识别日期格式：{value}")


def parse_yyyymmdd(value):
    return datetime.strptime(normalize_date_text(value), "%Y%m%d").date()


def date_range_desc(start_date, end_date):
    current = end_date
    while current >= start_date:
        yield current.strftime("%Y%m%d")
        current -= timedelta(days=1)


# ============================================================
# TDX 路径
# ============================================================


def is_tdx_root_dir(path):
    path = normalize_user_path(path, True)
    if not path or not os.path.isdir(path):
        return False

    return (os.path.isdir(win_join(path, "vipdoc"))
            and os.path.isdir(win_join(path, "T0002", "hq_cache")))


def resolve_tdx_root_dir(path):
    input_path = normalize_user_path(path, True)

    for folder in iter_parent_dirs(input_path):
        if is_tdx_root_dir(folder):
            return folder

    return input_path


def tdx_vipdoc_dir(tdx_root):
    return win_join(resolve_tdx_root_dir(tdx_root), "vipdoc")


def tdx_hq_cache_dir(tdx_root):
    return win_join(resolve_tdx_root_dir(tdx_root), "T0002", "hq_cache")


def infoharbor_block_path(tdx_root):
    return win_join(tdx_hq_cache_dir(tdx_root), "infoharbor_block.dat")


def tnf_path(tdx_root, market):
    return win_join(tdx_hq_cache_dir(tdx_root), TNF_FILES[market])


def day_file_path(tdx_root, market, symbol):
    prefix = MARKET_DAY_PREFIX[market]
    return win_join(tdx_vipdoc_dir(tdx_root), prefix, "lday",
                    f"{prefix}{symbol}.day")


# ============================================================
# THS 路径
# ============================================================


def is_ths_root_dir(path):
    path = normalize_user_path(path, True)
    if not path or not os.path.isdir(path):
        return False

    return (os.path.isfile(
        win_join(path, "BlockUpdate", THS_CONCEPT_BLOCK_FILE))
            and os.path.isdir(win_join(path, "stockname")))


def resolve_ths_root_dir(path):
    input_path = normalize_user_path(path, True)

    for folder in iter_parent_dirs(input_path):
        if is_ths_root_dir(folder):
            return folder

    return input_path


def ths_block_update_dir(ths_root):
    return win_join(resolve_ths_root_dir(ths_root), "BlockUpdate")


def ths_concept_block_path(ths_root):
    return win_join(ths_block_update_dir(ths_root), THS_CONCEPT_BLOCK_FILE)


def ths_stockname_dir(ths_root):
    return win_join(resolve_ths_root_dir(ths_root), "stockname")


# ============================================================
# 配置
# ============================================================


def normalize_excel_color(value):
    color = str(value or "").strip().replace("#", "").upper()

    if not re.fullmatch(r"[0-9A-F]{6}", color):
        return DEFAULT_MARK_FILL_COLOR

    return color


def make_default_config():
    start_date, end_date = default_start_end_dates()

    return {
        "config_version": 2,
        "ui": {
            "window_title": "概念板块统计排序助手 v1.1.0 20260525",
            "window_geometry": "1420x880",
        },
        "daily_source": {
            "type": DAILY_SOURCE_TDX,
            "tdx_root": "",
        },
        "concept_source": {
            "type": CONCEPT_SOURCE_TDX,
            "tdx_root": "",
            "ths_root": "",
        },
        "date_range": {
            "start_date": start_date,
            "end_date": end_date,
        },
        "entry_threshold": DEFAULT_ENTRY_THRESHOLD,
        "mark_threshold": DEFAULT_MARK_THRESHOLD,
        "mark_fill_color": DEFAULT_MARK_FILL_COLOR,
        "output": {
            "output_dir": "",
            "keep_all_blocks": False,
            "max_blocks_per_sheet": 20,
            "skip_zero_blocks": True,
            "auto_open_file": True,
        },
        "path_history": {
            "daily_tdx_root": "",
            "concept_tdx_root": "",
            "concept_ths_root": "",
            "output_dir": "",
        },
        "block_fields": {
            "order": [x["key"] for x in BLOCK_FIELD_DEFS],
            "selected": [x["key"] for x in BLOCK_FIELD_DEFS],
        },
        "detail_fields": {
            "order": [x["key"] for x in DETAIL_FIELD_DEFS],
            "selected": [x["key"] for x in DETAIL_FIELD_DEFS],
        },
    }


def deep_merge_known(default_obj, user_obj):
    result = deepcopy(default_obj)

    if not isinstance(user_obj, dict):
        return result

    for key in result.keys():
        if key not in user_obj:
            continue

        default_value = result[key]
        user_value = user_obj[key]

        if user_value is None:
            continue

        if isinstance(default_value, dict) and isinstance(user_value, dict):
            result[key] = deep_merge_known(default_value, user_value)
        else:
            result[key] = user_value

    return result


def normalize_config_paths(config):
    config["daily_source"]["tdx_root"] = normalize_user_path(
        config["daily_source"].get("tdx_root", ""), True)
    config["concept_source"]["tdx_root"] = normalize_user_path(
        config["concept_source"].get("tdx_root", ""), True)
    config["concept_source"]["ths_root"] = normalize_user_path(
        config["concept_source"].get("ths_root", ""), True)
    config["output"]["output_dir"] = normalize_user_path(
        config["output"].get("output_dir", ""),
        True,
    )
    config["mark_fill_color"] = normalize_excel_color(
        config.get("mark_fill_color", DEFAULT_MARK_FILL_COLOR))

    path_history = config.setdefault("path_history", {})
    for key in [
            "daily_tdx_root",
            "concept_tdx_root",
            "concept_ths_root",
            "output_dir",
    ]:
        path_history[key] = normalize_user_path(path_history.get(key, ""),
                                                True)

    return config


def load_app_config():
    default_config = make_default_config()
    path = get_config_path()

    if not os.path.exists(path):
        return normalize_config_paths(default_config)

    try:
        with open(path, "r", encoding="utf-8") as f:
            saved = json.load(f)
    except Exception:
        return normalize_config_paths(default_config)

    allowed = {
        "daily_source": saved.get("daily_source"),
        "concept_source": saved.get("concept_source"),
        "entry_threshold": saved.get("entry_threshold"),
        "mark_threshold": saved.get("mark_threshold"),
        "mark_fill_color": saved.get("mark_fill_color"),
        "output": saved.get("output"),
        "path_history": saved.get("path_history"),
        "block_fields": saved.get("block_fields"),
        "detail_fields": saved.get("detail_fields"),
    }

    config = deep_merge_known(default_config, allowed)

    start_date, end_date = default_start_end_dates()
    config["date_range"]["start_date"] = start_date
    config["date_range"]["end_date"] = end_date

    return normalize_config_paths(config)


def save_app_config(config):
    saved = {
        "config_version": 2,
        "daily_source": {
            "type":
            DAILY_SOURCE_TDX,
            "tdx_root":
            normalize_user_path(config["daily_source"]["tdx_root"], True),
        },
        "concept_source": {
            "type":
            config["concept_source"]["type"],
            "tdx_root":
            normalize_user_path(config["concept_source"]["tdx_root"], True),
            "ths_root":
            normalize_user_path(config["concept_source"]["ths_root"], True),
        },
        "entry_threshold": float(config["entry_threshold"]),
        "mark_threshold": float(config["mark_threshold"]),
        "mark_fill_color": normalize_excel_color(config["mark_fill_color"]),
        "path_history": {
            "daily_tdx_root":
            normalize_user_path(
                config.get("path_history", {}).get("daily_tdx_root", ""),
                True,
            ),
            "concept_tdx_root":
            normalize_user_path(
                config.get("path_history", {}).get("concept_tdx_root", ""),
                True,
            ),
            "concept_ths_root":
            normalize_user_path(
                config.get("path_history", {}).get("concept_ths_root", ""),
                True,
            ),
            "output_dir":
            normalize_user_path(
                config.get("path_history", {}).get("output_dir", ""),
                True,
            ),
        },
        "block_fields": deepcopy(config["block_fields"]),
        "detail_fields": deepcopy(config["detail_fields"]),
    }

    with open(get_config_path(), "w", encoding="utf-8") as f:
        json.dump(saved, f, ensure_ascii=False, indent=2)


# ============================================================
# TDX 名称解析
# ============================================================


def decode_ascii(raw):
    try:
        return raw.decode("ascii", errors="ignore").strip("\x00 ").strip()
    except Exception:
        return ""


def decode_gbk(raw):
    try:
        return raw.decode("gbk", errors="ignore").strip("\x00 ").strip()
    except Exception:
        return ""


def load_tnf_name_map_for_market(tdx_root, market):
    path = tnf_path(tdx_root, market)

    if not os.path.exists(path):
        return {}

    with open(path, "rb") as f:
        raw = f.read()

    if len(raw) <= TDX_TNF_HEADER_SIZE:
        return {}

    payload = raw[TDX_TNF_HEADER_SIZE:]
    total = len(payload) // TDX_TNF_RECORD_SIZE
    result = {}

    for idx in range(total):
        start = idx * TDX_TNF_RECORD_SIZE
        rec = payload[start:start + TDX_TNF_RECORD_SIZE]

        if len(rec) < TDX_TNF_RECORD_SIZE:
            continue

        symbol = decode_ascii(rec[0:20])
        name = decode_gbk(rec[31:63])

        if symbol:
            result[(market, symbol)] = name

    return result


def load_tdx_name_map(tdx_root, log_func):
    result = {}

    for market in ["SH", "SZ", "BJ"]:
        mp = load_tnf_name_map_for_market(tdx_root, market)
        result.update(mp)
        log_func(f"TDX 名称文件解析完成：{market}，数量：{len(mp)}")

    return result


# ============================================================
# TDX 概念成分解析
# ============================================================


def parse_infoharbor_gn_blocks(tdx_root):
    path = infoharbor_block_path(tdx_root)

    if not os.path.exists(path):
        raise FileNotFoundError(f"infoharbor_block.dat 不存在：\n{path}")

    with open(path, "r", encoding="gbk", errors="ignore", newline=None) as f:
        text = f.read()

    lines = text.splitlines()

    blocks = []
    current = None
    tokens = []

    def flush_current():
        nonlocal current, tokens

        if not current:
            return

        symbols = []
        seen = set()

        joined = ",".join(tokens)
        for item in joined.split(","):
            item = item.strip()
            if not item or "#" not in item:
                continue

            market_id, symbol = item.split("#", 1)
            market_id = market_id.strip()
            symbol = symbol.strip()

            market = MARKET_ID_MAP.get(market_id)
            if not market:
                continue

            if not re.fullmatch(r"\d{6}", symbol):
                continue

            key = (market, symbol)
            if key in seen:
                continue

            seen.add(key)
            symbols.append(key)

        current["symbols"] = symbols
        current["member_count_actual"] = len(symbols)

        if current["type"] == "GN":
            blocks.append(current)

        current = None
        tokens = []

    for line in lines:
        line = line.strip()

        if not line:
            continue

        if line.startswith("#"):
            flush_current()

            header = line[1:]
            parts = header.split(",")

            if not parts:
                continue

            name_part = parts[0]
            if "_" not in name_part:
                continue

            block_type, block_name = name_part.split("_", 1)

            member_count = 0
            if len(parts) > 1 and parts[1].strip().isdigit():
                member_count = int(parts[1].strip())

            block_code = parts[2].strip() if len(parts) > 2 else ""

            current = {
                "type": block_type.strip(),
                "name": block_name.strip(),
                "block_code": block_code,
                "symbols": [],
                "member_count": member_count,
            }
        else:
            if current:
                tokens.append(line)

    flush_current()
    return blocks


# ============================================================
# THS 文本解析
# ============================================================


def read_text_auto_encoding(path):
    encodings = ["gbk", "utf-8-sig", "utf-8", "ansi"]

    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return f.read()
        except LookupError:
            continue
        except Exception:
            continue

    with open(path, "r", encoding="gbk", errors="ignore") as f:
        return f.read()


def parse_ini_like_sections(path):
    text = read_text_auto_encoding(path)
    sections = {}
    current = None

    for raw_line in text.splitlines():
        line = raw_line.strip()

        if not line:
            continue

        if line.startswith(";") or line.startswith("#"):
            continue

        if line.startswith("[") and line.endswith("]"):
            current = line[1:-1].strip()
            sections.setdefault(current, {})
            continue

        if current and "=" in line:
            key, value = line.split("=", 1)
            sections[current][key.strip()] = value.strip()

    return sections


# ============================================================
# THS 概念成分解析
# ============================================================


def parse_ths_member_token(token):
    raw = str(token or "").strip()

    if not raw or ":" not in raw:
        return None

    if "*" in raw or "(" in raw or ")" in raw:
        return None

    prefix, symbol = raw.split(":", 1)
    prefix = prefix.strip()
    symbol = symbol.strip()

    prefix_abs = prefix.lstrip("-")
    market = THS_A_STOCK_PREFIX_TO_MARKET.get(prefix_abs)

    if not market:
        return None

    if not re.fullmatch(r"\d{6}", symbol):
        return None

    return market, symbol


def parse_ths_concept_blocks(ths_root):
    path = ths_concept_block_path(ths_root)

    if not os.path.exists(path):
        raise FileNotFoundError(f"同花顺概念成分文件不存在：\n{path}")

    sections = parse_ini_like_sections(path)

    name_map = sections.get("BLOCK_NAME_MAP_TABLE", {})
    context_map = sections.get("BLOCK_STOCK_CONTEXT", {})

    if not name_map:
        raise ValueError("同花顺 block_2B.ini 缺少 BLOCK_NAME_MAP_TABLE。")

    if not context_map:
        raise ValueError("同花顺 block_2B.ini 缺少 BLOCK_STOCK_CONTEXT。")

    blocks = []

    for block_id, block_name in name_map.items():
        raw_context = context_map.get(block_id, "")
        symbols = []
        seen = set()

        for token in raw_context.split(","):
            parsed = parse_ths_member_token(token)
            if not parsed:
                continue

            if parsed in seen:
                continue

            seen.add(parsed)
            symbols.append(parsed)

        if not symbols:
            continue

        blocks.append({
            "type": "GN",
            "name": block_name.strip(),
            "block_code": block_id.strip(),
            "symbols": symbols,
            "member_count": len(symbols),
            "member_count_actual": len(symbols),
        })

    blocks.sort(key=lambda x: x["name"])
    return blocks


# ============================================================
# THS 名称解析
# ============================================================


def collect_ths_concept_member_prefixes(ths_root):
    block_path = ths_concept_block_path(ths_root)
    sections = parse_ini_like_sections(block_path)

    context_map = sections.get("BLOCK_STOCK_CONTEXT", {})
    prefixes = set()

    for raw_context in context_map.values():
        for token in str(raw_context or "").split(","):
            token = token.strip()

            if not token or ":" not in token:
                continue

            prefix, _ = token.split(":", 1)
            prefix = prefix.strip().lstrip("-")

            if prefix:
                prefixes.add(prefix)

    return prefixes


def choose_ths_stockname_files(ths_root):
    stockname_dir = ths_stockname_dir(ths_root)
    prefixes = collect_ths_concept_member_prefixes(ths_root)

    market_ids = []
    seen = set()

    for prefix in sorted(prefixes):
        market_id = THS_BLOCK_PREFIX_TO_STOCKNAME_MARKET_ID.get(prefix)

        if not market_id:
            continue

        if market_id in seen:
            continue

        seen.add(market_id)
        market_ids.append(market_id)

    result = []

    for market_id in market_ids:
        path = win_join(stockname_dir, f"stockname_{market_id}_0.txt")

        if os.path.isfile(path):
            result.append(path)

    return result


def parse_ths_stockname_value(value):
    value = str(value or "").strip()

    if "|" not in value:
        return value.strip()

    short_name, _ = value.split("|", 1)
    return short_name.strip()


def load_ths_name_map(ths_root, log_func):
    stockname_dir = ths_stockname_dir(ths_root)

    if not os.path.isdir(stockname_dir):
        raise FileNotFoundError(f"同花顺 stockname 目录不存在：\n{stockname_dir}")

    paths = choose_ths_stockname_files(ths_root)
    result = {}

    if not paths:
        raise FileNotFoundError("未找到可用的同花顺当前主名称文件。\n\n"
                                "当前只允许读取 stockname_{MarketID}_0.txt，"
                                "不会读取 stockname_{MarketID}_1.txt 或其他历史名称文件。")

    for path in paths:
        sections = parse_ini_like_sections(path)

        for section, rows in sections.items():
            m = re.fullmatch(r"name_([^_]+)_([^_]+)", section)
            if not m:
                continue

            market_id = m.group(1)
            type_id = m.group(2)
            market = THS_NAME_SECTION_TO_MARKET.get((market_id, type_id))

            if not market:
                continue

            count = 0

            for code, value in rows.items():
                if code == "ConfigVer":
                    continue

                if not re.fullmatch(r"\d{6}", code):
                    continue

                name = parse_ths_stockname_value(value)
                if not name:
                    continue

                result[(market, code)] = name
                count += 1

            if count:
                log_func(f"THS 当前主名称解析完成：{section}，数量：{count}")

    log_func(f"THS 当前主名称文件解析完成：文件数：{len(paths)}，名称数：{len(result)}")
    return result


def filter_blacklisted_concept_blocks(blocks, log_func):
    result = []
    skipped = 0

    for block in blocks:
        block_name = str(block.get("name", "")).strip()

        if block_name in CONCEPT_BLOCK_NAME_BLACKLIST:
            skipped += 1
            continue

        result.append(block)

    if skipped:
        log_func(f"概念板块黑名单过滤：跳过 {skipped} 个")

    return result


# ============================================================
# 概念源统一入口
# ============================================================


def load_concept_blocks(config, log_func):
    source_type = config["concept_source"]["type"]

    if source_type == CONCEPT_SOURCE_TDX:
        tdx_root = resolve_tdx_root_dir(config["concept_source"]["tdx_root"])
        log_func("开始解析 TDX 概念成分表...")
        blocks = parse_infoharbor_gn_blocks(tdx_root)
        log_func(f"TDX 概念板块数量：{len(blocks)}")
        blocks = filter_blacklisted_concept_blocks(blocks, log_func)
        log_func(f"TDX 黑名单过滤后概念板块数量：{len(blocks)}")
        return blocks

    if source_type == CONCEPT_SOURCE_THS:
        ths_root = resolve_ths_root_dir(config["concept_source"]["ths_root"])
        log_func("开始解析 THS 概念成分表...")
        blocks = parse_ths_concept_blocks(ths_root)
        log_func(f"THS A股概念板块数量：{len(blocks)}")
        blocks = filter_blacklisted_concept_blocks(blocks, log_func)
        log_func(f"THS 黑名单过滤后概念板块数量：{len(blocks)}")
        return blocks

    raise ValueError(f"未知概念成分数据源：{source_type}")


def load_stock_name_map(config, log_func):
    source_type = config["concept_source"]["type"]

    if source_type == CONCEPT_SOURCE_TDX:
        tdx_root = resolve_tdx_root_dir(config["concept_source"]["tdx_root"])
        log_func("开始加载 TDX 股票名称...")
        return load_tdx_name_map(tdx_root, log_func)

    if source_type == CONCEPT_SOURCE_THS:
        ths_root = resolve_ths_root_dir(config["concept_source"]["ths_root"])
        log_func("开始加载 THS 股票名称...")
        return load_ths_name_map(ths_root, log_func)

    raise ValueError(f"未知概念成分数据源：{source_type}")


# ============================================================
# .day 倒序解析
# ============================================================


def unpack_day_record(rec):
    trade_date, open_i, high_i, low_i, close_i, amount_f, volume_i, _ = struct.unpack(
        "<IIIIIfII",
        rec,
    )

    return {
        "DATE": int(trade_date),
        "OPEN": float(open_i) / PRICE_DIVISOR,
        "HIGH": float(high_i) / PRICE_DIVISOR,
        "LOW": float(low_i) / PRICE_DIVISOR,
        "CLOSE": float(close_i) / PRICE_DIVISOR,
        "AMOUNT": float(amount_f),
        "VOLUME": int(volume_i),
    }


def parse_day_records_reverse_for_range(file_path, start_int, end_int):
    """
    倒序读取指定区间行情，并额外读取 start_int 之前最近一条记录用于计算首日涨幅。
    """
    if not os.path.exists(file_path):
        return []

    size = os.path.getsize(file_path)
    if size <= 0:
        return []

    if size % TDX_DAY_RECORD_SIZE != 0:
        return []

    total = size // TDX_DAY_RECORD_SIZE
    rows_desc = []

    with open(file_path, "rb") as f:
        for idx in range(total - 1, -1, -1):
            f.seek(idx * TDX_DAY_RECORD_SIZE)
            rec = f.read(TDX_DAY_RECORD_SIZE)

            if len(rec) != TDX_DAY_RECORD_SIZE:
                continue

            try:
                row = unpack_day_record(rec)
            except Exception:
                continue

            d = row["DATE"]

            if d > end_int:
                continue

            if d >= start_int:
                rows_desc.append(row)
                continue

            rows_desc.append(row)
            break

    if not rows_desc:
        return []

    return list(reversed(rows_desc))


def build_daily_stock_cache(tdx_root, symbols, name_map, start_date, end_date,
                            entry_threshold, log_func):
    start_int = int(start_date)
    end_int = int(end_date)

    daily_all = {}
    daily_qualified = {}

    symbols = sorted(symbols)
    total = len(symbols)

    valid_symbol_count = 0
    qualified_record_count = 0

    for idx, (market, symbol) in enumerate(symbols, start=1):
        path = day_file_path(tdx_root, market, symbol)
        rows = parse_day_records_reverse_for_range(path, start_int, end_int)

        if len(rows) < 2:
            if idx % 500 == 0:
                log_func(f"已解析日线：{idx}/{total}，有效证券：{valid_symbol_count}")
            continue

        valid_this_symbol = False

        for i in range(1, len(rows)):
            prev_row = rows[i - 1]
            row = rows[i]

            d = int(row["DATE"])
            if d < start_int or d > end_int:
                continue

            prev_close = float(prev_row["CLOSE"])
            if prev_close <= 0:
                continue

            pct_chg = (float(row["CLOSE"]) - prev_close) / prev_close * 100.0
            yyyymmdd = str(d)

            stock_data = {
                "MARKET": market,
                "SYMBOL": symbol,
                "SYMBOL_FULL": f"{market}{symbol}",
                "NAME": name_map.get((market, symbol), ""),
                "DATE": yyyymmdd,
                "OPEN": row["OPEN"],
                "HIGH": row["HIGH"],
                "LOW": row["LOW"],
                "CLOSE": row["CLOSE"],
                "VOLUME": row["VOLUME"],
                "AMOUNT": row["AMOUNT"],
                "PCT_CHG": pct_chg,
            }

            daily_all.setdefault(yyyymmdd, {})[(market, symbol)] = stock_data
            valid_this_symbol = True

            if pct_chg > entry_threshold:
                daily_qualified.setdefault(yyyymmdd,
                                           {})[(market, symbol)] = stock_data
                qualified_record_count += 1

        if valid_this_symbol:
            valid_symbol_count += 1

        if idx % 500 == 0:
            log_func(f"已解析日线：{idx}/{total}，有效证券：{valid_symbol_count}")

    log_func(f"日线解析完成：唯一证券 {total}，有效证券 {valid_symbol_count}，"
             f"达标证券日期记录 {qualified_record_count}")

    return daily_all, daily_qualified


# ============================================================
# 字段组件
# ============================================================


class HorizontalFieldGroup(tk.Frame):

    def __init__(self,
                 parent,
                 title,
                 field_defs,
                 order_keys,
                 selected_keys,
                 on_change=None):
        super().__init__(parent)

        self.title = title
        self.field_defs = field_defs
        self.field_map = {x["key"]: x for x in field_defs}
        self.valid_keys = [x["key"] for x in field_defs]

        self.order_keys = list(order_keys)
        self.selected_keys = set(selected_keys)
        self.items = []
        self.drag_key = None
        self.on_change = on_change

        self.normal_bg = "#ffffff"
        self.selected_bg = "#dcfce7"
        self.drag_bg = "#dbeafe"
        self.drop_bg = "#fff7cc"
        self.handle_normal_bg = "#eeeeee"
        self.handle_selected_bg = "#bbf7d0"

        self.validate_state()

        self.vars = {
            key: tk.BooleanVar(value=key in self.selected_keys)
            for key in self.valid_keys
        }

        self.build_ui()

    def validate_state(self):
        valid = set(self.valid_keys)

        if set(self.order_keys) != valid or len(self.order_keys) != len(
                self.valid_keys):
            self.order_keys = list(self.valid_keys)

        self.selected_keys = self.selected_keys & valid

    def notify_change(self):
        if self.on_change:
            self.on_change()

    def build_ui(self):
        toolbar = tk.Frame(self)
        toolbar.pack(fill="x", padx=3, pady=(1, 0))

        tk.Label(
            toolbar,
            text=f"{self.title}：勾选输出，拖动 ☰ 调整组内顺序",
            fg="#555555",
        ).pack(side="left")

        tk.Button(toolbar,
                  text="全选",
                  width=UI_BUTTON_WIDTH,
                  command=self.select_all).pack(side="right", padx=2)
        tk.Button(toolbar,
                  text="全不选",
                  width=UI_BUTTON_WIDTH,
                  command=self.unselect_all).pack(side="right", padx=2)

        container = tk.Frame(self)
        container.pack(fill="x", expand=False)

        self.canvas = tk.Canvas(container, height=34, highlightthickness=0)
        self.scrollbar = tk.Scrollbar(container,
                                      orient="horizontal",
                                      command=self.canvas.xview)
        self.canvas.configure(xscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="top", fill="x", expand=True)
        self.scrollbar.pack(side="bottom", fill="x")

        self.list_frame = tk.Frame(self.canvas)
        self.window_id = self.canvas.create_window((0, 0),
                                                   window=self.list_frame,
                                                   anchor="nw")

        self.list_frame.bind(
            "<Configure>", lambda e: self.canvas.configure(scrollregion=self.
                                                           canvas.bbox("all")))
        self.canvas.bind(
            "<Configure>",
            lambda e: self.canvas.itemconfigure(self.window_id,
                                                height=e.height))

        self.canvas.bind(
            "<Enter>",
            lambda e: self.canvas.bind_all("<MouseWheel>", self.on_mousewheel))
        self.canvas.bind("<Leave>",
                         lambda e: self.canvas.unbind_all("<MouseWheel>"))

        self.render_items()

    def on_mousewheel(self, event):
        self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

    def render_items(self):
        for child in self.list_frame.winfo_children():
            child.destroy()

        self.items = []

        for key in self.order_keys:
            title = self.field_map[key]["title"]

            item = tk.Frame(self.list_frame,
                            bd=1,
                            relief="solid",
                            bg=self.normal_bg)
            item.pack(side="left", fill="y", padx=1, pady=1)

            handle = tk.Label(item,
                              text="☰",
                              width=2,
                              cursor="fleur",
                              bg=self.handle_normal_bg)
            handle.pack(side="right", fill="y", padx=(1, 1), pady=0)

            cb = tk.Checkbutton(
                item,
                text=title,
                variable=self.vars[key],
                width=10,
                anchor="w",
                bg=self.normal_bg,
                command=self.refresh_colors,
            )
            cb.pack(side="left", fill="both", expand=True, padx=(2, 1), pady=0)

            handle.bind("<ButtonPress-1>",
                        lambda e, k=key: self.on_drag_start(k))
            handle.bind("<B1-Motion>", lambda e, k=key: self.on_drag_motion())
            handle.bind("<ButtonRelease-1>",
                        lambda e, k=key: self.on_drag_release())

            self.items.append({
                "key": key,
                "widget": item,
                "checkbox": cb,
                "handle": handle,
            })

        self.refresh_colors()

    def base_color(self, key):
        return self.selected_bg if self.vars[key].get() else self.normal_bg

    def handle_color(self, key):
        return self.handle_selected_bg if self.vars[key].get(
        ) else self.handle_normal_bg

    def refresh_colors(self):
        for item in self.items:
            key = item["key"]
            color = self.base_color(key)
            item["widget"].configure(bg=color)
            item["checkbox"].configure(bg=color,
                                       activebackground=color,
                                       selectcolor=color)
            item["handle"].configure(bg=self.handle_color(key))

        self.notify_change()

    def set_item_color(self, item, color, force_handle=False):
        item["widget"].configure(bg=color)
        item["checkbox"].configure(bg=color,
                                   activebackground=color,
                                   selectcolor=color)
        item["handle"].configure(
            bg=color if force_handle else self.handle_color(item["key"]))

    def on_drag_start(self, key):
        self.drag_key = key
        for item in self.items:
            if item["key"] == key:
                self.set_item_color(item, self.drag_bg, True)
            else:
                self.set_item_color(item, self.base_color(item["key"]))

    def calculate_drop_index(self):
        pointer_x = self.winfo_pointerx()

        for idx, item in enumerate(self.items):
            widget = item["widget"]
            mid = widget.winfo_rootx() + widget.winfo_width() / 2
            if pointer_x < mid:
                return idx

        return len(self.items)

    def on_drag_motion(self):
        if not self.drag_key:
            return

        drop_index = self.calculate_drop_index()

        for idx, item in enumerate(self.items):
            if idx == drop_index:
                self.set_item_color(item, self.drop_bg, True)
            elif item["key"] == self.drag_key:
                self.set_item_color(item, self.drag_bg, True)
            else:
                self.set_item_color(item, self.base_color(item["key"]))

    def on_drag_release(self):
        if not self.drag_key:
            return

        drop_index = self.calculate_drop_index()
        old_index = self.order_keys.index(self.drag_key)

        if drop_index > old_index:
            drop_index -= 1

        moved = self.order_keys.pop(old_index)
        self.order_keys.insert(drop_index, moved)

        self.drag_key = None
        self.render_items()
        self.notify_change()

    def select_all(self):
        for key in self.valid_keys:
            self.vars[key].set(True)
        self.refresh_colors()

    def unselect_all(self):
        for key in self.valid_keys:
            self.vars[key].set(False)
        self.refresh_colors()

    def get_order(self):
        return list(self.order_keys)

    def get_selected_in_order(self):
        return [key for key in self.order_keys if self.vars[key].get()]


# ============================================================
# 输出路径
# ============================================================


def default_output_path(config):
    start_str = config["date_range"]["start_date"]
    end_str = config["date_range"]["end_date"]
    filename = f"{start_str}-{end_str}_概念板块统计排序.xlsx"

    output_dir = normalize_user_path(config["output"]["output_dir"], True)
    if not output_dir:
        output_dir = get_program_dir()

    return win_join(output_dir, filename)


def make_unique_output_path(path):
    path = normalize_user_path(path, True)
    folder = win_dirname(path)
    filename = ntpath.basename(path)
    stem, ext = ntpath.splitext(filename)

    if not ext:
        ext = ".xlsx"

    idx = 1
    while True:
        candidate = win_join(folder, f"{stem}_{idx}{ext}")
        if not os.path.exists(candidate):
            return candidate
        idx += 1


# ============================================================
# 导出逻辑
# ============================================================


def make_title_map(field_defs):
    return {x["key"]: x["title"] for x in field_defs}


def build_rows_for_date(blocks, qualified_map, block_fields, detail_fields,
                        max_blocks, keep_all_blocks, skip_zero_blocks,
                        mark_threshold):
    block_title_map = make_title_map(BLOCK_FIELD_DEFS)
    detail_title_map = make_title_map(DETAIL_FIELD_DEFS)

    records = []

    for block in blocks:
        members = block["symbols"]
        member_count = int(block.get("member_count") or len(members) or 0)

        qualified = []
        for key in members:
            stock = qualified_map.get(key)
            if stock:
                qualified.append(stock)

        qualified.sort(
            key=lambda x: (-float(x["PCT_CHG"]), -float(x["AMOUNT"])))

        qualified_count = len(qualified)

        if skip_zero_blocks and qualified_count <= 0:
            continue

        ratio = qualified_count / member_count if member_count > 0 else 0

        records.append({
            "block": block,
            "qualified": qualified,
            "qualified_count": qualified_count,
            "ratio": ratio,
            "member_count": member_count,
        })

    records.sort(
        key=lambda x: (-x["qualified_count"], -x["ratio"], x["block"]["name"]))

    if not keep_all_blocks:
        records = records[:max_blocks]

    rows = []
    merge_ranges = []
    mark_rows = []
    current_excel_row = 2

    has_detail_fields = bool(detail_fields)

    appear_count_map = {}

    if has_detail_fields and "APPEAR_COUNT" in detail_fields:
        for rec in records:
            for stock in rec["qualified"]:
                symbol_full = stock.get("SYMBOL_FULL", "")
                if symbol_full:
                    appear_count_map[symbol_full] = appear_count_map.get(
                        symbol_full, 0) + 1

    for rank, rec in enumerate(records, start=1):
        block_values = {
            "RANK": rank,
            "BLOCK_NAME": rec["block"]["name"],
            "MEMBER_COUNT": rec["member_count"],
            "QUALIFIED_COUNT": rec["qualified_count"],
            "QUALIFIED_RATIO": rec["ratio"],
        }

        if not has_detail_fields:
            row = {}

            for key in block_fields:
                row[block_title_map[key]] = block_values.get(key, "")

            rows.append(row)
            current_excel_row += 1
            continue

        qualified = rec["qualified"] or [None]
        start_row = current_excel_row

        for stock in qualified:
            row = {}
            stock_values = stock or {}

            for key in block_fields:
                row[block_title_map[key]] = block_values.get(key, "")

            symbol_full = stock_values.get("SYMBOL_FULL", "")

            for key in detail_fields:
                if key == "APPEAR_COUNT":
                    value = appear_count_map.get(symbol_full, "")
                else:
                    value = stock_values.get(key, "")

                if key == "PCT_CHG" and value != "":
                    value = float(value) / 100.0

                row[detail_title_map[key]] = value

            rows.append(row)

            if stock and float(stock.get("PCT_CHG", 0)) > mark_threshold:
                mark_rows.append(current_excel_row)

            current_excel_row += 1

        end_row = current_excel_row - 1

        if end_row > start_row:
            merge_ranges.append((start_row, end_row))

    columns = [block_title_map[k] for k in block_fields
               ] + [detail_title_map[k] for k in detail_fields]

    return rows, columns, merge_ranges, len(records), records, mark_rows


def build_summary_rows(daily_rank_records, max_blocks):
    rank_limit = min(int(max_blocks), SUMMARY_MAX_RANK)
    name_col = "板块名称/排名"

    block_stat = {}

    for day_records in daily_rank_records:
        for rank, rec in enumerate(day_records[:rank_limit], start=1):
            block_name = rec["block"]["name"]

            if block_name not in block_stat:
                block_stat[block_name] = [0] * rank_limit

            # 注意：这里统计的是“排名前 x 的次数”
            for i in range(rank - 1, rank_limit):
                block_stat[block_name][i] += 1

    rows = []

    for block_name, counts in block_stat.items():
        row = {name_col: block_name}

        for i, count in enumerate(counts, start=1):
            row[str(i)] = count

        rows.append(row)

    def sort_key(row):
        counts = [int(row[str(i)]) for i in range(1, rank_limit + 1)]
        return tuple([-x for x in counts] + [row[name_col]])

    rows.sort(key=sort_key)

    columns = [name_col] + [str(i) for i in range(1, rank_limit + 1)]

    return rows, columns


def excel_display_text(value, number_format=None):
    if value is None:
        return ""

    if isinstance(value, float) or isinstance(value, int):
        if number_format == "0.00%":
            return f"{float(value) * 100:.2f}%"

        if number_format == "0.00":
            return f"{float(value):.2f}"

        if number_format == "#,##0.00":
            return f"{float(value):,.2f}"

        if number_format == "#,##0":
            return f"{int(value):,}"

    return str(value)


def text_visual_width(text):
    text = str(text or "")
    width = 0

    for ch in text:
        if ord(ch) > 127:
            width += 2
        else:
            width += 1

    return width


def style_sheet(writer, sheet_name, df, block_field_count, merge_ranges,
                mark_rows, mark_fill_color):
    ws = writer.sheets[sheet_name]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    mark_rows = set(mark_rows)
    mark_fill = PatternFill("solid", fgColor=mark_fill_color)

    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    if block_field_count > 0:
        for start_row, end_row in merge_ranges:
            for col_idx in range(1, block_field_count + 1):
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=end_row,
                    end_column=col_idx,
                )

    header_map = {cell.value: cell.column for cell in ws[1]}

    for title in ["达标占比"]:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = "0.00%"

    for title in ["当日涨幅"]:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = "0.00%"

    for title in ["开盘价", "最高价", "最低价", "收盘价"]:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = "0.00"

    for title in ["成交额"]:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = "#,##0.00"

    for title in ["成交量", "序号", "总成分股数", "达标个股数", "出现次数"]:
        col = header_map.get(title)
        if col:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = "#,##0"

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border

            if row_idx in mark_rows and col_idx > block_field_count:
                cell.fill = mark_fill

            if row_idx == 1:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                continue

            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header_value = ws.cell(1, col_idx).value

        max_width = text_visual_width(header_value)

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            display_text = excel_display_text(cell.value, cell.number_format)
            max_width = max(max_width, text_visual_width(display_text))

        width = max_width + 2
        width = max(8, min(width, 36))

        ws.column_dimensions[letter].width = width


def style_summary_sheet(ws):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, max_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left",
                                                         vertical="center")

    ws.freeze_panes = "B2"
    ws.sheet_view.selection[0].sqref = "A1"
    ws.sheet_view.selection[0].activeCell = "A1"

    ws.column_dimensions["A"].width = 24

    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4


def validate_tdx_daily_source(tdx_root):
    tdx_root = resolve_tdx_root_dir(tdx_root)

    if not tdx_root:
        raise ValueError("TDX 个股历史日线路径不能为空。")

    if not os.path.isdir(tdx_root):
        raise FileNotFoundError(f"TDX 个股历史日线目录不存在：\n{tdx_root}")

    vipdoc = tdx_vipdoc_dir(tdx_root)
    if not os.path.isdir(vipdoc):
        raise FileNotFoundError(f"TDX vipdoc 目录不存在：\n{vipdoc}")

    sh_lday = win_join(vipdoc, "sh", "lday")
    sz_lday = win_join(vipdoc, "sz", "lday")

    if not os.path.isdir(sh_lday):
        raise FileNotFoundError(f"TDX 沪市日线目录不存在：\n{sh_lday}")

    if not os.path.isdir(sz_lday):
        raise FileNotFoundError(f"TDX 深市日线目录不存在：\n{sz_lday}")

    return tdx_root


def validate_tdx_concept_source(tdx_root):
    tdx_root = resolve_tdx_root_dir(tdx_root)

    if not tdx_root:
        raise ValueError("TDX 概念成分表路径不能为空。")

    if not os.path.isdir(tdx_root):
        raise FileNotFoundError(f"TDX 概念成分表目录不存在：\n{tdx_root}")

    hq_cache = tdx_hq_cache_dir(tdx_root)
    if not os.path.isdir(hq_cache):
        raise FileNotFoundError(f"TDX T0002\\hq_cache 目录不存在：\n{hq_cache}")

    block_path = infoharbor_block_path(tdx_root)
    if not os.path.isfile(block_path):
        raise FileNotFoundError(f"TDX infoharbor_block.dat 不存在：\n{block_path}")

    return tdx_root


def validate_ths_concept_source(ths_root):
    ths_root = resolve_ths_root_dir(ths_root)

    if not ths_root:
        raise ValueError("THS 概念成分表路径不能为空。")

    if not os.path.isdir(ths_root):
        raise FileNotFoundError(f"THS 概念成分表目录不存在：\n{ths_root}")

    block_path = ths_concept_block_path(ths_root)
    if not os.path.isfile(block_path):
        raise FileNotFoundError(f"THS block_2B.ini 不存在：\n{block_path}")

    stockname_dir = ths_stockname_dir(ths_root)
    if not os.path.isdir(stockname_dir):
        raise FileNotFoundError(f"THS stockname 目录不存在：\n{stockname_dir}")

    return ths_root


def validate_config(config, output_path):
    daily_type = config["daily_source"]["type"]
    if daily_type != DAILY_SOURCE_TDX:
        raise ValueError(f"当前不支持的历史日线数据源：{daily_type}")

    daily_tdx_root = validate_tdx_daily_source(
        config["daily_source"]["tdx_root"])
    config["daily_source"]["tdx_root"] = daily_tdx_root

    concept_type = config["concept_source"]["type"]

    if concept_type == CONCEPT_SOURCE_TDX:
        concept_tdx_root = validate_tdx_concept_source(
            config["concept_source"]["tdx_root"])
        config["concept_source"]["tdx_root"] = concept_tdx_root
    elif concept_type == CONCEPT_SOURCE_THS:
        concept_ths_root = validate_ths_concept_source(
            config["concept_source"]["ths_root"])
        config["concept_source"]["ths_root"] = concept_ths_root
    else:
        raise ValueError(f"未知概念成分数据源：{concept_type}")

    start_date = parse_yyyymmdd(config["date_range"]["start_date"])
    end_date = parse_yyyymmdd(config["date_range"]["end_date"])

    if start_date > end_date:
        raise ValueError("开始日期不能晚于结束日期。")

    entry_threshold = float(config["entry_threshold"])
    if entry_threshold < -100:
        raise ValueError("入选涨幅阈值不合理。")

    mark_threshold = float(config["mark_threshold"])
    if mark_threshold < -100:
        raise ValueError("标记涨幅阈值不合理。")

    mark_fill_color = str(config["mark_fill_color"]).strip().replace(
        "#", "").upper()
    if not re.fullmatch(r"[0-9A-F]{6}", mark_fill_color):
        raise ValueError(f"标记颜色必须是 6 位十六进制颜色，例如 {DEFAULT_MARK_FILL_COLOR}。")

    max_blocks = int(config["output"]["max_blocks_per_sheet"])
    if max_blocks <= 0:
        raise ValueError("每日显示板块数必须是正整数。")

    output_path = normalize_user_path(output_path, True)

    if not output_path:
        raise ValueError("输出文件路径不能为空。")

    if not output_path.lower().endswith(".xlsx"):
        raise ValueError("输出文件必须是 .xlsx 格式。")

    output_dir = win_dirname(output_path)
    if output_dir and not os.path.isdir(output_dir):
        raise FileNotFoundError(f"输出目录不存在：\n{output_dir}")

    if not config["block_fields"]["selected"] and not config["detail_fields"][
            "selected"]:
        raise ValueError("至少需要选择一个输出字段。")


def export_to_excel(config, output_path, log_func=print):
    validate_config(config, output_path)

    daily_tdx_root = resolve_tdx_root_dir(config["daily_source"]["tdx_root"])
    start_date = normalize_date_text(config["date_range"]["start_date"])
    end_date = normalize_date_text(config["date_range"]["end_date"])
    entry_threshold = float(config["entry_threshold"])
    mark_threshold = float(config["mark_threshold"])
    mark_fill_color = normalize_excel_color(config["mark_fill_color"])

    block_fields = list(config["block_fields"]["selected"])
    detail_fields = list(config["detail_fields"]["selected"])

    keep_all_blocks = bool(config["output"]["keep_all_blocks"])
    max_blocks = int(config["output"]["max_blocks_per_sheet"])
    skip_zero_blocks = bool(config["output"]["skip_zero_blocks"])

    concept_type = config["concept_source"]["type"]

    log_func("历史日线数据源：TDX")
    log_func(f"TDX 日线路径：{daily_tdx_root}")
    log_func(f"概念成分数据源：{concept_type}")

    if concept_type == CONCEPT_SOURCE_TDX:
        log_func(
            f"TDX 概念路径：{resolve_tdx_root_dir(config['concept_source']['tdx_root'])}"
        )
    else:
        log_func(
            f"THS 概念路径：{resolve_ths_root_dir(config['concept_source']['ths_root'])}"
        )

    log_func(f"日期范围：{start_date} ~ {end_date}")
    log_func(f"入选涨幅阈值：>{entry_threshold}")
    log_func(f"标记涨幅阈值：>{mark_threshold}")
    log_func(f"标记颜色：#{mark_fill_color}")
    log_func(f"输出文件：{output_path}")

    blocks = load_concept_blocks(config, log_func)

    unique_symbols = set()
    for block in blocks:
        unique_symbols.update(block["symbols"])

    log_func(f"概念板块数量：{len(blocks)}")
    log_func(f"唯一证券数量：{len(unique_symbols)}")

    name_map = load_stock_name_map(config, log_func)

    log_func("开始倒序解析日线并计算涨幅...")
    daily_all, daily_qualified = build_daily_stock_cache(
        tdx_root=daily_tdx_root,
        symbols=unique_symbols,
        name_map=name_map,
        start_date=start_date,
        end_date=end_date,
        entry_threshold=entry_threshold,
        log_func=log_func,
    )

    daily_outputs = []
    daily_rank_records = []
    total_rows = 0

    for yyyymmdd in date_range_desc(parse_yyyymmdd(start_date),
                                    parse_yyyymmdd(end_date)):
        qualified_map = daily_qualified.get(yyyymmdd)

        if not qualified_map:
            log_func(f"处理日期：{yyyymmdd}，无达标个股，跳过")
            continue

        rows, columns, merge_ranges, block_count, rank_records, mark_rows = build_rows_for_date(
            blocks=blocks,
            qualified_map=qualified_map,
            block_fields=block_fields,
            detail_fields=detail_fields,
            max_blocks=max_blocks,
            keep_all_blocks=keep_all_blocks,
            skip_zero_blocks=skip_zero_blocks,
            mark_threshold=mark_threshold,
        )

        if not rows:
            log_func(f"处理日期：{yyyymmdd}，无达标板块，跳过")
            continue

        daily_outputs.append({
            "sheet_name": yyyymmdd,
            "rows": rows,
            "columns": columns,
            "merge_ranges": merge_ranges,
            "mark_rows": mark_rows,
            "block_count": block_count,
            "qualified_count": len(qualified_map),
            "rank_records": rank_records,
        })

        daily_rank_records.append(rank_records)
        total_rows += len(rows)

        log_func(f"处理日期：{yyyymmdd}，达标个股：{len(qualified_map)}，"
                 f"输出板块：{block_count}，输出行数：{len(rows)}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if daily_outputs:
            summary_sources = [
                ("近5日汇总", daily_rank_records[:5]),
                ("近10日汇总", daily_rank_records[:10]),
                ("全部汇总", daily_rank_records),
            ]

            for sheet_name, records in summary_sources:
                summary_rows, summary_columns = build_summary_rows(
                    records,
                    max_blocks,
                )

                summary_df = pd.DataFrame(summary_rows,
                                          columns=summary_columns)
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)

                summary_ws = writer.sheets[sheet_name]
                style_summary_sheet(summary_ws)

            for item in daily_outputs:
                df = pd.DataFrame(item["rows"], columns=item["columns"])
                df.to_excel(writer, sheet_name=item["sheet_name"], index=False)

                style_sheet(
                    writer=writer,
                    sheet_name=item["sheet_name"],
                    df=df,
                    block_field_count=len(block_fields),
                    merge_ranges=item["merge_ranges"],
                    mark_rows=item["mark_rows"],
                    mark_fill_color=mark_fill_color,
                )

            first_ws = writer.sheets["近5日汇总"]
            writer.book.active = writer.book.sheetnames.index("近5日汇总")
            first_ws.sheet_view.selection[0].sqref = "A1"
            first_ws.sheet_view.selection[0].activeCell = "A1"

        else:
            info_df = pd.DataFrame({
                "说明": [
                    "指定日期范围内没有生成任何达标板块数据。",
                    f"日期范围：{start_date} ~ {end_date}",
                    f"入选涨幅阈值：>{entry_threshold}",
                    f"标记涨幅阈值：>{mark_threshold}",
                ]
            })

            info_df.to_excel(writer, sheet_name="近5日汇总", index=False)

            summary_ws = writer.sheets["近5日汇总"]
            style_summary_sheet(summary_ws)

            writer.book.active = writer.book.sheetnames.index("近5日汇总")
            summary_ws.sheet_view.selection[0].sqref = "A1"
            summary_ws.sheet_view.selection[0].activeCell = "A1"

    log_func("-" * 50)
    log_func("导出完成。")
    log_func(f"汇总 Sheet 数量：{3 if daily_outputs else 1}")
    log_func(f"每日明细 Sheet 数量：{len(daily_outputs)}")
    log_func(f"总行数：{total_rows}")
    log_func(f"文件位置：{output_path}")


# ============================================================
# GUI
# ============================================================


class App:

    def __init__(self, root):
        self.root = root
        self.config = load_app_config()
        self.log_queue = queue.Queue()
        self.auto_save_job = None
        self.is_building_ui = True

        self.root.title(self.config["ui"]["window_title"])
        self.set_window_icon()
        self.root.geometry(self.config["ui"]["window_geometry"])

        try:
            self.root.state("zoomed")
        except Exception:
            pass

        self.build_vars()
        self.build_ui()
        self.update_output_path_by_dates()
        self.bind_auto_save_events()
        self.is_building_ui = False

        self.root.after(100, self.poll_log_queue)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def build_vars(self):
        self.daily_tdx_root_var = tk.StringVar(
            value=self.config["daily_source"]["tdx_root"])

        self.concept_source_type_var = tk.StringVar(
            value=self.config["concept_source"]["type"])
        self.concept_tdx_root_var = tk.StringVar(
            value=self.config["concept_source"]["tdx_root"])
        self.concept_ths_root_var = tk.StringVar(
            value=self.config["concept_source"]["ths_root"])

        self.start_date_var = tk.StringVar(
            value=self.config["date_range"]["start_date"])
        self.end_date_var = tk.StringVar(
            value=self.config["date_range"]["end_date"])

        self.entry_threshold_var = tk.StringVar(
            value=str(self.config["entry_threshold"]))
        self.mark_threshold_var = tk.StringVar(
            value=str(self.config["mark_threshold"]))
        self.mark_fill_color_var = tk.StringVar(
            value=str(self.config["mark_fill_color"]))

        self.output_path_var = tk.StringVar(
            value=default_output_path(self.config))

        self.keep_all_blocks_var = tk.BooleanVar(
            value=self.config["output"]["keep_all_blocks"])
        self.max_blocks_var = tk.StringVar(
            value=str(self.config["output"]["max_blocks_per_sheet"]))
        self.skip_zero_blocks_var = tk.BooleanVar(
            value=self.config["output"]["skip_zero_blocks"])
        self.auto_open_file_var = tk.BooleanVar(
            value=self.config["output"].get("auto_open_file", True))

    def build_ui(self):
        main = tk.Frame(self.root)
        main.pack(fill="both", expand=True, padx=10, pady=6)

        title_frame = tk.Frame(main)
        title_frame.pack(fill="x", pady=(0, 4))

        tk.Label(
            title_frame,
            text="概念板块统计排序导出为多页Excel表——汇总表及每日明细表",
            font=("Microsoft YaHei", 15, "bold"),
        ).pack(side="left", anchor="w")

        tk.Button(
            title_frame,
            text="反馈交流",
            width=10,
            command=self.show_feedback_qr,
        ).pack(side="right")

        self.build_data_source_frame(main)
        self.build_param_frame(main)
        self.build_output_frame(main)
        self.build_field_frame(main)
        self.build_log_frame(main)

    def build_data_source_frame(self, parent):
        wrapper = tk.LabelFrame(parent, text="数据源")
        wrapper.pack(fill="x", pady=3)

        self.build_daily_source_frame(wrapper)
        self.build_concept_source_frame(wrapper)

    def build_daily_source_frame(self, parent):
        frame = tk.LabelFrame(parent, text="个股历史日线数据源")
        frame.pack(fill="x", padx=6, pady=(4, 2))

        tk.Label(frame, text="TDX 个股历史日线：").grid(row=0,
                                                 column=0,
                                                 sticky="e",
                                                 padx=8,
                                                 pady=6)

        tk.Entry(frame, textvariable=self.daily_tdx_root_var,
                 width=120).grid(row=0, column=1, sticky="we", padx=6, pady=6)

        tk.Button(frame,
                  text="选择文件夹",
                  width=UI_BUTTON_WIDTH,
                  command=self.choose_daily_tdx_root).grid(row=0,
                                                           column=2,
                                                           sticky="w",
                                                           padx=4,
                                                           pady=6)

        tk.Button(frame,
                  text="打开文件夹",
                  width=UI_BUTTON_WIDTH,
                  command=lambda: self.open_existing_folder(
                      self.daily_tdx_root_var.get())).grid(row=0,
                                                           column=3,
                                                           sticky="w",
                                                           padx=4,
                                                           pady=6)

        frame.columnconfigure(1, weight=1)

    def build_concept_source_frame(self, parent):
        frame = tk.LabelFrame(parent, text="概念板块成分数据源")
        frame.pack(fill="x", padx=6, pady=(2, 4))

        tk.Radiobutton(
            frame,
            text="TDX 概念成分表",
            variable=self.concept_source_type_var,
            value=CONCEPT_SOURCE_TDX,
            command=self.refresh_concept_source_state,
        ).grid(row=0, column=0, sticky="w", padx=8, pady=6)

        self.concept_tdx_entry = tk.Entry(
            frame, textvariable=self.concept_tdx_root_var, width=120)
        self.concept_tdx_entry.grid(row=0,
                                    column=1,
                                    sticky="we",
                                    padx=6,
                                    pady=6)

        self.concept_tdx_choose_btn = tk.Button(
            frame,
            text="选择文件夹",
            width=UI_BUTTON_WIDTH,
            command=self.choose_concept_tdx_root)

        self.concept_tdx_choose_btn.grid(row=0,
                                         column=2,
                                         sticky="w",
                                         padx=4,
                                         pady=6)

        self.concept_tdx_open_btn = tk.Button(
            frame,
            text="打开文件夹",
            width=UI_BUTTON_WIDTH,
            command=lambda: self.open_existing_folder(self.concept_tdx_root_var
                                                      .get()))
        self.concept_tdx_open_btn.grid(row=0,
                                       column=3,
                                       sticky="w",
                                       padx=4,
                                       pady=6)

        tk.Radiobutton(
            frame,
            text="THS 概念成分表",
            variable=self.concept_source_type_var,
            value=CONCEPT_SOURCE_THS,
            command=self.refresh_concept_source_state,
        ).grid(row=1, column=0, sticky="w", padx=8, pady=6)

        self.concept_ths_entry = tk.Entry(
            frame, textvariable=self.concept_ths_root_var, width=120)
        self.concept_ths_entry.grid(row=1,
                                    column=1,
                                    sticky="we",
                                    padx=6,
                                    pady=6)

        self.concept_ths_choose_btn = tk.Button(
            frame,
            text="选择文件夹",
            width=UI_BUTTON_WIDTH,
            command=self.choose_concept_ths_root)
        self.concept_ths_choose_btn.grid(row=1,
                                         column=2,
                                         sticky="w",
                                         padx=4,
                                         pady=6)

        self.concept_ths_open_btn = tk.Button(
            frame,
            text="打开文件夹",
            width=UI_BUTTON_WIDTH,
            command=lambda: self.open_existing_folder(self.concept_ths_root_var
                                                      .get()))
        self.concept_ths_open_btn.grid(row=1,
                                       column=3,
                                       sticky="w",
                                       padx=4,
                                       pady=6)

        frame.columnconfigure(1, weight=1)
        self.refresh_concept_source_state()

    def build_param_frame(self, parent):
        frame = tk.LabelFrame(parent, text="日期与统计参数")
        frame.pack(fill="x", pady=3)

        tk.Label(frame, text="开始：").grid(row=0,
                                         column=0,
                                         sticky="e",
                                         padx=(6, 2),
                                         pady=6)
        start_entry = tk.Entry(frame,
                               textvariable=self.start_date_var,
                               width=max(10,
                                         len(self.start_date_var.get()) + 1))
        start_entry.grid(row=0, column=1, sticky="w", padx=(2, 6), pady=6)
        start_entry.bind(
            "<FocusOut>",
            lambda e: self.normalize_date_entry(self.start_date_var))

        tk.Label(frame, text="结束：").grid(row=0,
                                         column=2,
                                         sticky="e",
                                         padx=(6, 2),
                                         pady=6)
        end_entry = tk.Entry(frame,
                             textvariable=self.end_date_var,
                             width=max(10,
                                       len(self.end_date_var.get()) + 1))
        end_entry.grid(row=0, column=3, sticky="w", padx=(2, 6), pady=6)
        end_entry.bind("<FocusOut>",
                       lambda e: self.normalize_date_entry(self.end_date_var))

        tk.Label(frame, text="入选涨幅 >").grid(row=0,
                                            column=4,
                                            sticky="e",
                                            padx=(6, 2),
                                            pady=6)
        tk.Entry(
            frame,
            textvariable=self.entry_threshold_var,
            width=max(5,
                      len(self.entry_threshold_var.get()) + 2),
        ).grid(row=0, column=5, sticky="w", padx=(2, 1), pady=6)
        tk.Label(frame, text="%").grid(row=0,
                                       column=6,
                                       sticky="w",
                                       padx=(0, 6),
                                       pady=6)

        tk.Label(frame, text="标记涨幅 >").grid(row=0,
                                            column=7,
                                            sticky="e",
                                            padx=(6, 2),
                                            pady=6)
        tk.Entry(
            frame,
            textvariable=self.mark_threshold_var,
            width=max(5,
                      len(self.mark_threshold_var.get()) + 2),
        ).grid(row=0, column=8, sticky="w", padx=(2, 1), pady=6)
        tk.Label(frame, text="%").grid(row=0,
                                       column=9,
                                       sticky="w",
                                       padx=(0, 6),
                                       pady=6)

        tk.Label(frame, text="颜色：").grid(row=0,
                                         column=10,
                                         sticky="e",
                                         padx=(6, 2),
                                         pady=6)

        self.mark_color_preview = tk.Label(
            frame,
            width=4,
            relief="solid",
            bd=1,
            bg=f"#{normalize_excel_color(self.mark_fill_color_var.get())}",
        )
        self.mark_color_preview.grid(row=0,
                                     column=11,
                                     sticky="w",
                                     padx=(2, 3),
                                     pady=6)

        tk.Button(
            frame,
            text="选色",
            width=6,
            command=self.choose_mark_color,
        ).grid(row=0, column=12, sticky="w", padx=(0, 8), pady=6)

        tk.Label(frame, text="板块数：").grid(row=0,
                                          column=13,
                                          sticky="e",
                                          padx=(6, 2),
                                          pady=6)
        self.max_blocks_entry = tk.Entry(
            frame,
            textvariable=self.max_blocks_var,
            width=max(5,
                      len(self.max_blocks_var.get()) + 2),
        )
        self.max_blocks_entry.grid(row=0,
                                   column=14,
                                   sticky="w",
                                   padx=(2, 6),
                                   pady=6)

        tk.Checkbutton(
            frame,
            text="全部保留",
            variable=self.keep_all_blocks_var,
            command=self.toggle_max_blocks_entry,
        ).grid(row=0, column=15, sticky="w", padx=(2, 6), pady=6)

        tk.Checkbutton(
            frame,
            text="不输出达标个股数为0的板块",
            variable=self.skip_zero_blocks_var,
        ).grid(row=0, column=16, sticky="w", padx=(2, 6), pady=6)

        self.start_date_var.trace_add(
            "write", lambda *_: self.update_output_path_by_dates())
        self.end_date_var.trace_add(
            "write", lambda *_: self.update_output_path_by_dates())

        self.mark_fill_color_var.trace_add(
            "write", lambda *_: self.refresh_mark_color_preview())

        self.toggle_max_blocks_entry()

    def build_output_frame(self, parent):
        frame = tk.LabelFrame(parent, text="输出")
        frame.pack(fill="x", pady=3)

        tk.Label(frame, text="输出文件：").grid(row=0,
                                           column=0,
                                           sticky="e",
                                           padx=8,
                                           pady=6)

        tk.Entry(frame, textvariable=self.output_path_var,
                 width=120).grid(row=0, column=1, sticky="we", padx=6, pady=6)

        tk.Button(frame,
                  text="选择文件夹",
                  width=UI_BUTTON_WIDTH,
                  command=self.choose_output_dir).grid(row=0,
                                                       column=2,
                                                       sticky="w",
                                                       padx=4,
                                                       pady=6)

        tk.Button(frame,
                  text="另存为",
                  width=UI_BUTTON_WIDTH,
                  command=self.choose_output_file).grid(row=0,
                                                        column=3,
                                                        sticky="w",
                                                        padx=4,
                                                        pady=6)

        tk.Button(frame,
                  text="打开文件夹",
                  width=UI_BUTTON_WIDTH,
                  command=self.open_output_folder).grid(row=0,
                                                        column=4,
                                                        sticky="w",
                                                        padx=4,
                                                        pady=6)

        frame.columnconfigure(1, weight=1)

    def build_field_frame(self, parent):
        frame = tk.LabelFrame(parent, text="输出字段")
        frame.pack(fill="x", pady=3)

        self.block_field_group = HorizontalFieldGroup(
            frame,
            title="板块字段",
            field_defs=BLOCK_FIELD_DEFS,
            order_keys=self.config["block_fields"]["order"],
            selected_keys=self.config["block_fields"]["selected"],
            on_change=self.schedule_auto_save,
        )
        self.block_field_group.pack(fill="x", expand=False, padx=4, pady=1)

        self.detail_field_group = HorizontalFieldGroup(
            frame,
            title="达标个股明细字段",
            field_defs=DETAIL_FIELD_DEFS,
            order_keys=self.config["detail_fields"]["order"],
            selected_keys=self.config["detail_fields"]["selected"],
            on_change=self.schedule_auto_save,
        )
        self.detail_field_group.pack(fill="x", expand=False, padx=4, pady=1)

    def build_log_frame(self, parent):
        frame = tk.LabelFrame(parent, text="日志")
        frame.pack(fill="both", expand=True, pady=3)

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        text_container = tk.Frame(frame)
        text_container.grid(row=0,
                            column=0,
                            sticky="nsew",
                            padx=6,
                            pady=(3, 2))
        text_container.rowconfigure(0, weight=1)
        text_container.columnconfigure(0, weight=1)

        self.log_text = tk.Text(text_container, height=12, wrap="none")
        self.log_text.grid(row=0, column=0, sticky="nsew")

        y_scroll = tk.Scrollbar(text_container,
                                orient="vertical",
                                command=self.log_text.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")

        x_scroll = tk.Scrollbar(text_container,
                                orient="horizontal",
                                command=self.log_text.xview)
        x_scroll.grid(row=1, column=0, sticky="we")

        self.log_text.configure(yscrollcommand=y_scroll.set,
                                xscrollcommand=x_scroll.set)

        btn_frame = tk.Frame(frame)
        btn_frame.grid(row=1, column=0, sticky="e", padx=4, pady=(1, 3))

        tk.Checkbutton(
            btn_frame,
            text="自动打开成果文件",
            variable=self.auto_open_file_var,
        ).pack(side="left", padx=8)

        self.export_btn = tk.Button(btn_frame,
                                    text="开始统计",
                                    width=UI_MAIN_BUTTON_WIDTH,
                                    command=self.start_export)
        self.export_btn.pack(side="left", padx=4)

    def set_window_icon(self):
        icon_path = get_resource_path(ICON_FILE_NAME)

        if not os.path.exists(icon_path):
            return

        try:
            self.root.iconbitmap(icon_path)
        except Exception:
            pass

    def bind_popup_escape_close(self, popup):
        popup.bind("<Escape>", lambda event: popup.destroy())
        popup.protocol("WM_DELETE_WINDOW", popup.destroy)

    def show_feedback_qr(self):
        qr_path = get_resource_path(QR_CODE_FILE_NAME)

        if not os.path.exists(qr_path):
            messagebox.showwarning(
                "二维码不存在",
                "未找到微信二维码图片：\n\n"
                f"{qr_path}\n\n"
                "请将图片命名为 wechat_qr.png，并放在程序目录下。",
            )
            return

        popup = tk.Toplevel(self.root)
        popup.title("反馈交流")
        popup.resizable(False, False)
        self.bind_popup_escape_close(popup)

        tk.Label(
            popup,
            text="扫码添加作者微信，反馈交流",
            font=("Microsoft YaHei", 12, "bold"),
        ).pack(padx=18, pady=(16, 8))

        try:
            original_image = tk.PhotoImage(file=qr_path)
        except Exception as e:
            messagebox.showerror(
                "图片加载失败",
                f"二维码图片加载失败：\n\n{qr_path}\n\n原因：{e}\n\n"
                "建议使用 PNG 格式图片。",
            )
            popup.destroy()
            return

        max_qr_size = 360
        image_width = original_image.width()
        image_height = original_image.height()
        max_side = max(image_width, image_height)
        scale = max(1, (max_side + max_qr_size - 1) // max_qr_size)

        qr_image = original_image.subsample(scale, scale)

        image_label = tk.Label(popup, image=qr_image)
        image_label.image = qr_image
        image_label.pack(padx=18, pady=8)

        tk.Button(
            popup,
            text="关闭",
            width=10,
            command=popup.destroy,
        ).pack(pady=(4, 16))

        popup.transient(self.root)
        popup.grab_set()
        popup.update_idletasks()

        x = self.root.winfo_rootx() + (self.root.winfo_width() -
                                       popup.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() -
                                       popup.winfo_height()) // 2

        popup.geometry(f"+{x}+{y}")
        popup.focus_set()

    def refresh_concept_source_state(self):
        source_type = self.concept_source_type_var.get()

        tdx_state = "normal" if source_type == CONCEPT_SOURCE_TDX else "disabled"
        ths_state = "normal" if source_type == CONCEPT_SOURCE_THS else "disabled"

        for widget in [
                self.concept_tdx_entry,
                self.concept_tdx_choose_btn,
                self.concept_tdx_open_btn,
        ]:
            widget.config(state=tdx_state)

        for widget in [
                self.concept_ths_entry,
                self.concept_ths_choose_btn,
                self.concept_ths_open_btn,
        ]:
            widget.config(state=ths_state)

        self.schedule_auto_save()

    def get_dialog_initial_dir(self, current_path, history_key):
        current_path = normalize_user_path(current_path, True)

        if current_path:
            if os.path.isfile(current_path):
                current_dir = win_dirname(current_path)
            else:
                current_dir = current_path

            if os.path.isdir(current_dir):
                return current_dir

        history = self.config.get("path_history", {}).get(history_key, "")
        history = normalize_user_path(history, True)

        if history and os.path.isdir(history):
            return history

        return get_program_dir()

    def update_path_history(self, history_key, path):
        path = normalize_user_path(path, True)

        if not path:
            return

        if os.path.isfile(path):
            path = win_dirname(path)

        if not os.path.isdir(path):
            return

        self.config.setdefault("path_history", {})[history_key] = path

    def ask_directory_for_path(self, title, current_path, history_key):
        initial_dir = self.get_dialog_initial_dir(current_path, history_key)

        path = filedialog.askdirectory(
            title=title,
            initialdir=initial_dir,
        )

        if not path:
            return ""

        path = normalize_user_path(path, True)
        self.update_path_history(history_key, path)

        return path

    def choose_daily_tdx_root(self):
        path = self.ask_directory_for_path(
            title="选择 TDX 个股历史日线目录",
            current_path=self.daily_tdx_root_var.get(),
            history_key="daily_tdx_root",
        )

        if path:
            resolved = resolve_tdx_root_dir(path)
            self.daily_tdx_root_var.set(resolved)
            self.update_path_history("daily_tdx_root", resolved)
            self.schedule_auto_save()

    def choose_concept_tdx_root(self):
        path = self.ask_directory_for_path(
            title="选择 TDX 概念成分表目录",
            current_path=self.concept_tdx_root_var.get(),
            history_key="concept_tdx_root",
        )

        if path:
            resolved = resolve_tdx_root_dir(path)
            self.concept_tdx_root_var.set(resolved)
            self.update_path_history("concept_tdx_root", resolved)
            self.schedule_auto_save()

    def choose_concept_ths_root(self):
        path = self.ask_directory_for_path(
            title="选择 THS 概念成分表目录",
            current_path=self.concept_ths_root_var.get(),
            history_key="concept_ths_root",
        )

        if path:
            resolved = resolve_ths_root_dir(path)
            self.concept_ths_root_var.set(resolved)
            self.update_path_history("concept_ths_root", resolved)
            self.schedule_auto_save()

    def choose_output_dir(self):
        path = self.ask_directory_for_path(
            title="选择输出目录",
            current_path=win_dirname(self.output_path_var.get()),
            history_key="output_dir",
        )

        if path:
            path = normalize_user_path(path, True)
            self.config["output"]["output_dir"] = path
            self.update_path_history("output_dir", path)
            self.update_output_path_by_dates(force_dir=path)
            self.schedule_auto_save()

    def choose_output_file(self):
        try:
            start_str = normalize_date_text(self.start_date_var.get())
            end_str = normalize_date_text(self.end_date_var.get())
        except Exception:
            start_str = self.start_date_var.get().strip()
            end_str = self.end_date_var.get().strip()

        initial_dir = self.get_dialog_initial_dir(
            current_path=self.output_path_var.get(),
            history_key="output_dir",
        )

        path = filedialog.asksaveasfilename(
            title="选择输出 Excel 文件",
            defaultextension=".xlsx",
            initialdir=initial_dir,
            initialfile=f"{start_str}-{end_str}_概念板块统计排序.xlsx",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("所有文件", "*.*"),
            ],
        )

        if path:
            path = normalize_user_path(path, True)
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.output_path_var.set(path)
            output_dir = normalize_user_path(win_dirname(path), True)
            self.config["output"]["output_dir"] = output_dir
            self.update_path_history("output_dir", output_dir)
            self.schedule_auto_save()

    def open_output_folder(self):
        output_path = normalize_user_path(self.output_path_var.get(), True)
        folder = win_dirname(
            output_path
        ) if output_path else self.config["output"]["output_dir"]
        self.open_existing_folder(folder)

    def open_existing_folder(self, folder):
        folder = normalize_user_path(folder, True)

        if not folder:
            messagebox.showwarning("无法打开", "路径为空。")
            return

        if not os.path.isdir(folder):
            messagebox.showwarning("无法打开", f"文件夹不存在：\n{folder}")
            return

        try:
            os.startfile(folder)
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件夹：\n{folder}\n\n原因：{e}")

    def normalize_date_entry(self, var):
        raw = var.get().strip()
        if not raw:
            return

        try:
            var.set(normalize_date_text(raw))
            self.update_output_path_by_dates()
        except Exception:
            pass

    def refresh_mark_color_preview(self):
        if not hasattr(self, "mark_color_preview"):
            return

        try:
            color = normalize_excel_color(self.mark_fill_color_var.get())
        except Exception:
            color = DEFAULT_MARK_FILL_COLOR

        self.mark_color_preview.config(bg=f"#{color}")

    def choose_mark_color(self):
        current = normalize_excel_color(self.mark_fill_color_var.get())

        result = colorchooser.askcolor(
            color=f"#{current}",
            title="选择标记颜色",
        )

        if not result or not result[1]:
            return

        color = result[1].replace("#", "").upper()
        self.mark_fill_color_var.set(color)
        self.refresh_mark_color_preview()
        self.schedule_auto_save()

    def update_output_path_by_dates(self, force_dir=None):
        try:
            start_str = normalize_date_text(self.start_date_var.get())
            end_str = normalize_date_text(self.end_date_var.get())
        except Exception:
            return

        output_dir = force_dir if force_dir is not None else self.config[
            "output"]["output_dir"]
        output_dir = normalize_user_path(output_dir, True)

        if not output_dir:
            output_dir = get_program_dir()

        self.output_path_var.set(
            win_join(output_dir, f"{start_str}-{end_str}_概念板块统计排序.xlsx"))

    def toggle_max_blocks_entry(self):
        if self.keep_all_blocks_var.get():
            self.max_blocks_entry.config(state="disabled")
        else:
            self.max_blocks_entry.config(state="normal")

        self.schedule_auto_save()

    def collect_config_from_ui(self):
        start = normalize_date_text(self.start_date_var.get())
        end = normalize_date_text(self.end_date_var.get())

        if parse_yyyymmdd(start) > parse_yyyymmdd(end):
            raise ValueError("开始日期不能晚于结束日期。")

        try:
            entry_threshold = float(self.entry_threshold_var.get().strip())
        except Exception:
            raise ValueError("入选涨幅阈值必须是数字。")

        try:
            mark_threshold = float(self.mark_threshold_var.get().strip())
        except Exception:
            raise ValueError("标记涨幅阈值必须是数字。")

        mark_fill_color = self.mark_fill_color_var.get().strip().replace(
            "#", "").upper()
        if not re.fullmatch(r"[0-9A-F]{6}", mark_fill_color):
            raise ValueError(
                f"标记颜色必须是 6 位十六进制颜色，例如 {DEFAULT_MARK_FILL_COLOR}。")

        self.mark_fill_color_var.set(mark_fill_color)

        raw_max = self.max_blocks_var.get().strip()
        if not raw_max.isdigit() or int(raw_max) <= 0:
            raise ValueError("每日显示板块数必须是正整数。")

        config = deepcopy(self.config)

        config.setdefault("path_history",
                          deepcopy(self.config.get("path_history", {})))
        config["daily_source"]["type"] = DAILY_SOURCE_TDX
        config["daily_source"]["tdx_root"] = resolve_tdx_root_dir(
            self.daily_tdx_root_var.get())

        concept_type = self.concept_source_type_var.get()
        if concept_type not in [CONCEPT_SOURCE_TDX, CONCEPT_SOURCE_THS]:
            raise ValueError("请选择概念板块成分数据源。")

        config["concept_source"]["type"] = concept_type
        config["concept_source"]["tdx_root"] = resolve_tdx_root_dir(
            self.concept_tdx_root_var.get())
        config["concept_source"]["ths_root"] = resolve_ths_root_dir(
            self.concept_ths_root_var.get())

        config["date_range"]["start_date"] = start
        config["date_range"]["end_date"] = end
        config["entry_threshold"] = entry_threshold
        config["mark_threshold"] = mark_threshold
        config["mark_fill_color"] = mark_fill_color

        output_path = normalize_user_path(self.output_path_var.get(), True)
        if output_path and not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"
            self.output_path_var.set(output_path)

        config["output"]["output_dir"] = normalize_user_path(
            win_dirname(output_path), True)
        config["output"]["keep_all_blocks"] = bool(
            self.keep_all_blocks_var.get())
        config["output"]["max_blocks_per_sheet"] = int(raw_max)
        config["output"]["skip_zero_blocks"] = bool(
            self.skip_zero_blocks_var.get())
        config["output"]["auto_open_file"] = bool(
            self.auto_open_file_var.get())

        config["block_fields"]["order"] = self.block_field_group.get_order()
        config["block_fields"][
            "selected"] = self.block_field_group.get_selected_in_order()

        config["detail_fields"]["order"] = self.detail_field_group.get_order()
        config["detail_fields"][
            "selected"] = self.detail_field_group.get_selected_in_order()

        if not config["block_fields"]["selected"] and not config[
                "detail_fields"]["selected"]:
            raise ValueError("至少需要选择一个输出字段。")

        return normalize_config_paths(config)

    def resolve_existing_output_file(self, output_path):
        output_path = normalize_user_path(output_path, True)

        if not os.path.exists(output_path):
            return output_path

        choice = messagebox.askyesnocancel(
            "文件已存在",
            "输出文件已存在：\n\n"
            f"{output_path}\n\n"
            "请选择处理方式：\n\n"
            "是：覆盖原文件\n"
            "否：自动改名保存\n"
            "取消：取消本次导出",
        )

        if choice is True:
            return output_path

        if choice is False:
            new_path = make_unique_output_path(output_path)
            self.output_path_var.set(new_path)
            return new_path

        return None

    def can_write_output_file(self, output_path):
        output_path = normalize_user_path(output_path, True)

        try:
            folder = win_dirname(output_path)
            if folder and not os.path.isdir(folder):
                return False

            if os.path.exists(output_path):
                with open(output_path, "a+b"):
                    pass
            else:
                with open(output_path, "wb"):
                    pass
                os.remove(output_path)

            return True
        except Exception:
            return False

    def resolve_writable_output_file(self, output_path):
        current_path = normalize_user_path(output_path, True)

        while True:
            if self.can_write_output_file(current_path):
                return current_path

            choice = messagebox.askyesnocancel(
                "文件无法写入",
                "当前输出文件无法写入，可能正在被 Excel 或其他程序打开：\n\n"
                f"{current_path}\n\n"
                "请选择处理方式：\n\n"
                "已经关闭文件选【是】\n"
                "要改名另存选【否】\n"
                "中止本次导出选【取消】",
            )

            if choice is True:
                continue

            if choice is False:
                current_path = make_unique_output_path(current_path)
                self.output_path_var.set(current_path)
                continue

            return None

    def bind_auto_save_events(self):
        vars_to_watch = [
            self.daily_tdx_root_var,
            self.concept_source_type_var,
            self.concept_tdx_root_var,
            self.concept_ths_root_var,
            self.start_date_var,
            self.end_date_var,
            self.entry_threshold_var,
            self.mark_threshold_var,
            self.mark_fill_color_var,
            self.output_path_var,
            self.keep_all_blocks_var,
            self.max_blocks_var,
            self.skip_zero_blocks_var,
            self.auto_open_file_var,
        ]

        for var in vars_to_watch:
            var.trace_add("write", lambda *_: self.schedule_auto_save())

    def schedule_auto_save(self):
        if getattr(self, "is_building_ui", False):
            return

        if self.auto_save_job is not None:
            self.root.after_cancel(self.auto_save_job)

        self.auto_save_job = self.root.after(500,
                                             self.auto_save_current_config)

    def auto_save_current_config(self):
        self.auto_save_job = None

        try:
            self.config = self.collect_config_from_ui()
            save_app_config(self.config)
        except Exception:
            pass

    def log(self, msg):
        self.log_queue.put(str(msg))

    def poll_log_queue(self):
        while True:
            try:
                msg = self.log_queue.get_nowait()
            except queue.Empty:
                break

            self.log_text.insert("end", msg + "\n")
            self.log_text.see("end")

        self.root.after(100, self.poll_log_queue)

    def start_export(self):
        try:
            config = self.collect_config_from_ui()
            output_path = normalize_user_path(self.output_path_var.get(), True)

            if not output_path.lower().endswith(".xlsx"):
                output_path += ".xlsx"
                self.output_path_var.set(output_path)

            validate_config(config, output_path)

            resolved = self.resolve_existing_output_file(output_path)
            if resolved is None:
                return

            writable_path = self.resolve_writable_output_file(resolved)
            if writable_path is None:
                return

            output_path = normalize_user_path(writable_path, True)
            self.output_path_var.set(output_path)

            config["output"]["output_dir"] = normalize_user_path(
                win_dirname(output_path), True)

            self.config = config
            save_app_config(self.config)

        except Exception as e:
            messagebox.showerror("输入错误", str(e))
            return

        self.export_btn.config(state="disabled")
        self.log_text.delete("1.0", "end")

        thread = threading.Thread(
            target=self.run_export_thread,
            args=(deepcopy(self.config), output_path),
            daemon=True,
        )
        thread.start()

    def run_export_thread(self, config, output_path):
        try:
            export_to_excel(config, output_path, log_func=self.log)
            self.log("任务成功完成。")

            auto_open = bool(config["output"].get("auto_open_file", True))

            def after_success():
                if auto_open:
                    try:
                        os.startfile(output_path)
                    except Exception as e:
                        self.log(f"文件已生成，但自动打开失败：{e}")
                        messagebox.showinfo("完成", "导出完成，但自动打开文件失败。")
                        return

                    messagebox.showinfo("完成", "导出完成，已自动打开文件。")
                else:
                    messagebox.showinfo("完成", "导出完成。")

            self.root.after(0, after_success)

        except Exception as e:
            msg = str(e)
            self.log(f"导出失败：{msg}")
            self.root.after(0, lambda m=msg: messagebox.showerror("导出失败", m))

        finally:
            self.root.after(0, lambda: self.export_btn.config(state="normal"))

    def on_close(self):
        try:
            self.config = self.collect_config_from_ui()
            save_app_config(self.config)
        except Exception:
            pass

        self.root.destroy()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()

# 以下内容要始终保留，不得删除
# 【运行虚拟环境】
# PS E:\AppProject\GNBlockRank> .venv\\Scripts\activate

# 【封装exe文件】
# Set-Location "E:\AppProject\GNBlockRank"
# pyinstaller "E:\AppProject\GNBlockRank\GNBlockRank.py" --onefile --windowed --clean --noconfirm --name "GNBlockRank" --icon "E:\AppProject\GNBlockRank\icon.ico" --add-data "E:\AppProject\GNBlockRank\wechat_qr.png;." --add-data "E:\AppProject\GNBlockRank\icon.ico;." --upx-dir "D:\upx-5.1.1-win64" --hidden-import secrets --exclude-module matplotlib --exclude-module scipy --exclude-module PyQt5 --exclude-module PyQt6 --exclude-module PySide2 --exclude-module PySide6 --hidden-import openpyxl.styles --exclude-module IPython --exclude-module notebook --exclude-module pytest --exclude-module unittest --exclude-module pydoc --exclude-module doctest --exclude-module html --exclude-module http --exclude-module xmlrpc

# 文件会生成在"E:\AppProject\GNBlockRank\dist\GNBlockRank.exe"

# 【github版本管理】

# Set-Location "E:\AppProject\GNBlockRank"
# git init #初始化，仅最初运行一次，后续不再运行

# git add . #把所有代码加入暂存区，需要每次运行
# git commit -m "提交说明" #提交代码到本地仓库

# git remote add origin https://github.com/mccoach/gnblockrank.git #关联远程仓库（最关键一步），仅第一次推送前运行

# git push origin main #推送代码到远程仓库，只推代码不推版本号

# git tag 版本号（不能含空格） #提交版本号，不能包含空格，如v1.0.0.20260524

# git push origin --tags #推送所有版本号标签到远程仓库，只推版本号不推代码

# #最常用的 5 条命令（每次推送都用）
# git add .            # 保存改动
# git commit -m "说明"  # 提交到本地
# git push origin main # 推送到远程
# git tag 版本号      # 打版本号标签
# git push origin --tags  # 推送版本号
